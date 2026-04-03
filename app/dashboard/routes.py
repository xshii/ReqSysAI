from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from flask import abort, current_app, flash, g, jsonify, redirect, render_template, request, send_file, session, url_for
from app.utils.api import api_ok, api_err
from flask_login import current_user, login_required
from sqlalchemy.orm import joinedload

from app.constants import REQ_INACTIVE_STATUSES, TODO_STATUS_DONE
from app.dashboard import dashboard_bp
from app.decorators import manager_required
from app.extensions import db
from app.models.project import Project
from app.models.project_member import ProjectMember
from app.models.report import PersonalWeekly, WeeklyReport
from app.models.requirement import Requirement
from app.models.risk import Risk
from app.models.todo import Todo, todo_requirements
from app.models.user import Group, Role, User
from app.services.ai import call_ollama
from app.services.prompts import get_prompt
from app.services.statistics import (
    gather_week_stats,
    get_delivery_metrics,
    get_estimate_deviation,
    get_reviewer,
    get_todo_progress,
    week_range,
)


def _merge_member_roles(members):
    """Merge roles for same person across projects, deduplicated."""
    from collections import OrderedDict
    result = OrderedDict()
    for m in members:
        name = m.display_name
        if name not in result:
            result[name] = set()
        result[name].add(m.role_label)
    return {name: '/'.join(sorted(roles)) for name, roles in result.items()}


def _build_people_tree(project_id, sub_project_ids):
    """Build tree: project → project_role(去括号聚类) → [{name, note}].

    project_role 中的括号内容提取为人员的独有职责(note)。
    """
    import re
    from collections import OrderedDict
    all_pids = [project_id] + sub_project_ids
    members = ProjectMember.query.filter(
        ProjectMember.project_id.in_(all_pids)
    ).options(
        db.joinedload(ProjectMember.project),
        db.joinedload(ProjectMember.user),
    ).order_by(ProjectMember.sort_order).all()

    seen_names = set()
    _seen_entries = set()  # (proj_name, role_group, name) for O(1) dedup
    tree = OrderedDict()  # project_name → {role_group → [person]}
    for m in members:
        proj_name = m.project.name if m.project else '未分配'
        name = m.display_name

        # Parse project_role: strip parentheses for grouping, extract note
        raw_role = (m.project_role or '').strip()
        match = re.match(r'^([^(（]+?)(?:\s*[(（](.+?)[)）])?\s*$', raw_role)
        if match:
            role_group = match.group(1).strip() or 'DEV'
            note = (match.group(2) or '').strip()
        else:
            role_group = raw_role or 'DEV'
            note = ''

        seen_names.add(name)

        if proj_name not in tree:
            tree[proj_name] = OrderedDict()
        if role_group not in tree[proj_name]:
            tree[proj_name][role_group] = []
        _key = (proj_name, role_group, name)
        if _key not in _seen_entries:
            _seen_entries.add(_key)
            tree[proj_name][role_group].append({'name': name, 'note': note})

    # Ensure parent project appears first in the tree
    parent_proj = Project.query.get(project_id)
    parent_name = parent_proj.name if parent_proj else None
    if parent_name and parent_name in tree:
        reordered = OrderedDict()
        reordered[parent_name] = tree[parent_name]
        for k, v in tree.items():
            if k != parent_name:
                reordered[k] = v
        tree = reordered

    unique_count = len(seen_names)
    tree._unique_count = unique_count
    return tree


def _gen_people_tree_img(project_id, sub_project_ids, project_name=''):
    """Generate people tree chart image (base64 PNG)."""
    try:
        from app.services.people_tree import generate_people_tree_image
        tree = _build_people_tree(project_id, sub_project_ids)
        if tree:
            return generate_people_tree_image(tree, project_name=project_name)
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning('People tree image failed: %s', e)
    return None


def _urgency_sort(reqs, limit=None):
    """Sort requirements: overdue → active(urgency desc) → done(ahead days desc)."""
    today_ = date.today()

    def _key(r):
        is_done = r.status in ('done', 'closed')
        is_overdue = (r.due_date and r.due_date < today_ and not is_done)
        group = 0 if is_overdue else (2 if is_done else 1)
        pct = 100 if is_done else (r.completion or 0)
        remain = 100 - pct
        days_left = max((r.due_date - today_).days, 1) if r.due_date else 999
        urgency = -(remain / days_left)  # negate for asc sort
        ahead = 0
        if is_done and r.due_date and r.updated_at:
            ahead = -(r.due_date - r.updated_at.date()).days
        return (group, urgency, ahead)

    result = sorted(reqs, key=_key)
    return result[:limit] if limit else result


def _guard_hidden_project(cur_project_id):
    """If project_id refers to a hidden project and user is not manager, return (None, None)."""
    if not cur_project_id:
        return cur_project_id, None
    p = db.session.get(Project, cur_project_id)
    if p and p.id in g.hidden_pids:
        return None, None
    return cur_project_id, p


def _visible_projects():
    """Active projects visible to current user."""
    _hset = set(g.hidden_pids)
    return [p for p in Project.query.filter_by(status='active').order_by(Project.name).all()
            if p.id not in _hset]


def _build_sub_projects(cur_project, monday):
    """Build sub-project progress list for parent project weekly report."""
    if not cur_project or not cur_project.children:
        return []
    sub_projects = []
    for child in cur_project.children:
        child_saved = WeeklyReport.query.filter_by(
            project_id=child.id, week_start=monday).first()
        from app.models.project_member import ProjectMember as PM_
        pm = PM_.query.filter_by(project_id=child.id, project_role='PM').first()
        fo = pm or PM_.query.filter_by(project_id=child.id, project_role='FO').first()
        summary = child_saved.summary if child_saved and child_saved.summary else None
        if summary is None:
            # AI generate one-line summary for child project (single aggregate query)
            from sqlalchemy import func as _fn, case as _case
            _today_d = date.today()
            stats = db.session.query(
                _fn.count(Requirement.id),
                _fn.sum(_case((Requirement.status.in_(('done', 'closed')), 1), else_=0)),
                _fn.sum(_case((Requirement.status == 'in_progress', 1), else_=0)),
                _fn.sum(_case((
                    db.and_(
                        Requirement.due_date < _today_d,
                        Requirement.status.notin_(('done', 'closed')),
                        Requirement.due_date.isnot(None),
                    ), 1), else_=0)),
            ).filter(
                Requirement.project_id == child.id,
                Requirement.parent_id.is_(None),
            ).first()
            c_total = stats[0] or 0
            c_done = int(stats[1] or 0)
            c_dev = int(stats[2] or 0)
            c_overdue = int(stats[3] or 0)
            # Completed todos this week (subquery for req ids)
            from app.models.todo import todo_requirements as tr_
            child_req_subq = db.session.query(Requirement.id).filter(
                Requirement.project_id == child.id,
                Requirement.parent_id.is_(None),
            ).subquery()
            week_done = Todo.query.filter(
                Todo.done_date >= monday, Todo.done_date <= _today_d
            ).join(tr_, Todo.id == tr_.c.todo_id).filter(
                tr_.c.requirement_id.in_(db.session.query(child_req_subq.c.id))
            ).count()
            context = (f'{child.name}：需求 {c_done}/{c_total} 完成，{c_dev}个开发中，'
                       f'本周完成 {week_done} 个todo'
                       + (f'，{c_overdue}个延期' if c_overdue else ''))
            try:
                _, raw = call_ollama(f'用一句话（不超过30字）总结以下项目进展，直接输出文字：\n{context}')
                summary = (raw or '').strip()[:50] if raw else context
            except Exception:
                summary = context
        sub_projects.append({
            'project': child,
            'owner': fo.user if fo and fo.user else (child.owner if child.owner else None),
            'summary': summary,
        })
    return sub_projects


@dashboard_bp.route('/requirements')
@login_required
def requirement_progress():

    cur_status = request.args.get('status', '')
    cur_project_id = request.args.get('project_id', type=int)
    cur_project_id, _ = _guard_hidden_project(cur_project_id)

    query = Requirement.query.options(
        joinedload(Requirement.project), joinedload(Requirement.assignee),
    )
    if cur_status:
        query = query.filter_by(status=cur_status)
    if cur_project_id:
        query = query.filter_by(project_id=cur_project_id)
    if g.hidden_pids:
        query = query.filter(Requirement.project_id.notin_(g.hidden_pids))
    requirements = query.order_by(Requirement.updated_at.desc()).all()

    todo_counts = get_todo_progress([r.id for r in requirements])

    return render_template('dashboard/requirements.html',
        requirements=requirements, todo_counts=todo_counts,
        projects=_visible_projects(),
        statuses=Requirement.STATUS_LABELS,
        cur_status=cur_status, cur_project=cur_project_id,
        today=date.today(),
    )



def _get_pinned_knowledge(project_id, sub_project_ids=None):
    """Get pinned knowledge items for project (+ children)."""
    from app.models.knowledge import Knowledge
    if not project_id:
        return []
    pids = [project_id] + (sub_project_ids or [])
    return Knowledge.query.filter(
        Knowledge.project_id.in_(pids), Knowledge.is_pinned == True
    ).order_by(Knowledge.project_id, Knowledge.title).all()


# ---- Weekly delta helper ----

def _compute_weekly_deltas(all_reqs, project_ids, monday, sunday):
    """Compute week-over-week delta indicators from existing data."""
    from app.models.requirement import Activity

    today_ = date.today()
    req_ids = {r.id for r in all_reqs}

    # 1. 本周完成的需求数 (Activity记录 或 updated_at本周+status=done)
    done_activities = Activity.query.filter(
        Activity.action == 'status_changed',
        Activity.detail.contains('→ 已完成'),
        Activity.created_at >= monday,
        Activity.created_at <= sunday + timedelta(days=1),
    ).all()
    done_req_ids = {a.requirement_id for a in done_activities if a.requirement_id in req_ids}
    # Fallback: 没有Activity记录但updated_at在本周且status=done的也算
    for r in all_reqs:
        if r.id not in done_req_ids and r.status in ('done', 'closed') \
                and r.updated_at and r.updated_at.date() >= monday and r.updated_at.date() <= sunday:
            done_req_ids.add(r.id)
    done_delta = len(done_req_ids)
    tc_done_delta = 0
    for r in all_reqs:
        if r.id in done_req_ids and r.source == 'testing' and r.test_cases:
            tc_done_delta += r.test_cases

    # 3. 延期变化
    overdue_count = sum(1 for r in all_reqs
                        if r.due_date and r.due_date < today_
                        and r.status not in ('done', 'closed'))
    new_overdue = sum(1 for r in all_reqs
                      if r.due_date and monday <= r.due_date < today_
                      and r.status not in ('done', 'closed'))
    resolved_overdue = sum(1 for r in all_reqs
                          if r.id in done_req_ids and r.due_date and r.due_date < monday)
    overdue_delta = new_overdue - resolved_overdue

    # 4. 按时交付率: 已完成需求中，完成时间 <= due_date 的占比
    done_reqs = [r for r in all_reqs if r.status in ('done', 'closed')]
    on_time = sum(1 for r in done_reqs
                  if r.due_date and r.updated_at and r.updated_at.date() <= r.due_date)
    on_time_total = sum(1 for r in done_reqs if r.due_date)
    on_time_rate = round(on_time / on_time_total * 100) if on_time_total else None

    # 5. 加权完成率变化（估算）
    _active = [r for r in all_reqs if r.status != 'closed']
    _comp_pct = lambda r: 100 if r.status == 'done' else (r.completion or 0)
    _comp_w_sum = sum(_comp_pct(r) * (r.estimate_days or 1) for r in _active)
    _comp_d_sum = sum((r.estimate_days or 1) for r in _active)
    current_weighted = round(_comp_w_sum / _comp_d_sum) if _comp_d_sum else 0

    if done_req_ids and _comp_d_sum:
        est_prev_sum = _comp_w_sum
        for r in all_reqs:
            if r.id in done_req_ids and r.status in ('done', 'closed'):
                days = r.estimate_days or 1
                est_prev_sum -= 100 * days
                est_prev_sum += 60 * days
        prev_weighted = round(est_prev_sum / _comp_d_sum)
        completion_delta = current_weighted - prev_weighted
    else:
        completion_delta = 0

    # 6. 本周新增需求数
    new_req_delta = sum(1 for r in all_reqs
                        if r.created_at and r.created_at.date() >= monday
                        and r.created_at.date() <= sunday)

    # 7. 本周完成的人天数
    done_days_delta = sum(r.estimate_days or 0 for r in all_reqs if r.id in done_req_ids)

    # 7. AI辅助人天 = sum(estimate_days * ai_ratio / 100)
    ai_days = sum((r.estimate_days or 0) * (r.ai_ratio or 0) / 100 for r in all_reqs)
    ai_days_delta = sum((r.estimate_days or 0) * (r.ai_ratio or 0) / 100
                        for r in all_reqs if r.id in done_req_ids)
    est_total = sum(r.estimate_days or 0 for r in all_reqs)

    # 8. 用例基线: 总用例数 + 已完成用例数(按加权完成率向下取整)
    import math as _math
    tc_total_all = 0
    tc_done_all = 0
    for r in all_reqs:
        if r.source == 'testing' and r.test_cases and r.test_cases > 0:
            tc_total_all += r.test_cases
            wpct = 100 if r.status in ('done', 'closed') else r.weighted_completion
            tc_done_all += _math.floor(r.test_cases * wpct / 100)

    return {
        'done_delta': done_delta,
        'done_days_delta': round(done_days_delta, 1),
        'tc_done_delta': tc_done_delta,
        'overdue_count': overdue_count,
        'overdue_delta': overdue_delta,
        'on_time': on_time,
        'on_time_total': on_time_total,
        'on_time_rate': on_time_rate,
        'completion_delta': completion_delta,
        'ai_days': round(ai_days, 1),
        'ai_days_delta': round(ai_days_delta, 1),
        'new_req_delta': new_req_delta,
        'est_total': round(est_total, 1),
        'tc_total_all': tc_total_all,
        'tc_done_all': tc_done_all,
    }


# ---- Rule-based weekly summary (no AI) ----

def _build_rule_based_summary(project_name, all_reqs, todos_done, todos_active,
                              open_risks, milestones, monday, sunday):
    """Generate a rule-based weekly report summary when AI service is disabled."""
    today_ = date.today()

    # Summary
    total = len(all_reqs)
    done = sum(1 for r in all_reqs if r.status in ('done', 'closed'))
    overdue = sum(1 for r in all_reqs if r.due_date and r.due_date < today_ and r.status not in ('done', 'closed'))
    active = total - done
    pct = round(done / total * 100) if total else 0
    summary = f'{project_name}本周（{monday.strftime("%m/%d")}~{sunday.strftime("%m/%d")}）：需求 {done}/{total} 完成（{pct}%）'
    if overdue:
        summary += f'，{overdue} 项延期'
    summary += f'；本周完成 {len(todos_done)} 个任务，进行中 {len(todos_active)} 个'
    if open_risks:
        summary += f'；{len(open_risks)} 个风险待处理'
    summary += '。'

    # Highlights: completed requirements this week + milestone progress
    highlights = []
    for r in all_reqs:
        if r.status in ('done', 'closed') and r.updated_at and r.updated_at.date() >= monday:
            highlights.append(f'需求 [{r.number}] {r.title} 已完成')
    for m in (milestones or []):
        if m.status == 'completed':
            highlights.append(f'里程碑「{m.name}」已达成')
    if not highlights:
        if todos_done:
            highlights.append(f'本周完成 {len(todos_done)} 个任务')

    # Risks: from open_risks
    risk_lines = []
    for r in open_risks:
        line = f'{r.title}（{r.severity_label}）'
        if r.due_date and r.due_date < today_:
            line += f' - 已超期{(today_ - r.due_date).days}天'
        risk_lines.append(line)

    # Plan: overdue + active items
    plan = []
    for r in all_reqs:
        if r.due_date and r.due_date < today_ and r.status not in ('done', 'closed'):
            plan.append(f'推进延期需求 [{r.number}] {r.title}')
    for m in (milestones or []):
        if m.status != 'completed' and m.due_date:
            days_left = (m.due_date - today_).days
            if 0 <= days_left <= 14:
                plan.append(f'里程碑「{m.name}」将于 {m.due_date.strftime("%m-%d")} 到期')
    if not plan:
        plan.append('继续推进各项需求')

    return {
        'summary': summary,
        'highlights': highlights[:5],
        'risks': risk_lines[:5],
        'plan': plan[:5],
    }


# ---- Pivot / 点灯图 helper ----

def _build_pivot_data(project_id, include_sub=True):
    """Build pivot table data for category L1 x L2 x source.
    Returns a dict of template variables, or empty dict if no pivot data."""
    import math
    from collections import Counter, defaultdict

    _parent_ids = {r.parent_id for r in db.session.query(Requirement.parent_id).filter(Requirement.parent_id.isnot(None)).distinct()}
    pivot_query = Requirement.query.filter(
        Requirement.category.isnot(None), Requirement.category.contains('-'),
        Requirement.id.notin_(_parent_ids) if _parent_ids else db.true())
    if project_id:
        if include_sub:
            child_ids = [c.id for c in Project.query.filter_by(parent_id=project_id).all()]
            pivot_query = pivot_query.filter(Requirement.project_id.in_([project_id] + child_ids))
        else:
            pivot_query = pivot_query.filter(Requirement.project_id == project_id)
    if g.hidden_pids:
        pivot_query = pivot_query.filter(Requirement.project_id.notin_(g.hidden_pids))
    pivot_reqs = pivot_query.all()
    if not pivot_reqs:
        return {}

    _all_sources = ['analysis', 'coding', 'testing']
    _src_labels = Requirement.SOURCE_LABELS
    today_ = date.today()

    _l1_min_start, _l2_min_start = {}, {}
    _l1_max_due, _l2_max_due = {}, {}
    for r in pivot_reqs:
        l1, l2 = r.category_l1, r.category_l2
        if r.start_date:
            if l1 not in _l1_min_start or r.start_date < _l1_min_start[l1]:
                _l1_min_start[l1] = r.start_date
            if l2 not in _l2_min_start or r.start_date < _l2_min_start[l2]:
                _l2_min_start[l2] = r.start_date
        if r.due_date:
            if l1 not in _l1_max_due or r.due_date > _l1_max_due[l1]:
                _l1_max_due[l1] = r.due_date
            if l2 not in _l2_max_due or r.due_date > _l2_max_due[l2]:
                _l2_max_due[l2] = r.due_date
    _far_future = date(2099, 1, 1)
    pivot_l1s = sorted(set(r.category_l1 for r in pivot_reqs), key=lambda x: (_l1_min_start.get(x, _far_future), _l1_max_due.get(x, _far_future)))
    pivot_l2s = sorted(set(r.category_l2 for r in pivot_reqs), key=lambda x: (_l2_min_start.get(x, _far_future), _l2_max_due.get(x, _far_future)))

    def _empty_cell():
        return {'total': 0, 'days': 0, 'done': 0, 'overdue': 0, 'active': 0, 'not_started': 0,
                'wpct_sum': 0, 'assignee': '', 'light': 'secondary',
                'tc_total': 0, 'tc_done': 0,
                'min_start': None, 'max_due': None,
                'overdue_assignee': '', 'overdue_days': 0}

    def _classify(r):
        if r.status in ('done', 'closed'):
            return 'done'
        if r.due_date and r.due_date < today_ and r.status not in ('done', 'closed'):
            return 'overdue'
        if r.status in ('pending', 'pending_review') and not (r.completion and r.completion > 0):
            return 'not_started'
        return 'active'

    def _finish_cell(c, names):
        if c['overdue'] > 0:
            c['light'] = 'danger'
        elif c['done'] == c['total'] and c['total'] > 0:
            c['light'] = 'success'
        elif c['active'] > 0:
            c['light'] = 'primary'
        else:
            c['light'] = 'secondary'
        if names:
            c['assignee'] = Counter(names).most_common(1)[0][0]

    pivot_src_cells = defaultdict(_empty_cell)
    _src_assignees = defaultdict(list)

    for r in pivot_reqs:
        src = r.source or 'coding'
        key = (r.category_l1, r.category_l2, src)
        c = pivot_src_cells[key]
        c['total'] += 1
        c['days'] += r.estimate_days or 0
        c['wpct_sum'] += r.weighted_completion
        cls = _classify(r)
        c[cls] += 1
        if cls == 'overdue' and r.due_date:
            od = (today_ - r.due_date).days
            if od > c['overdue_days']:
                c['overdue_days'] = od
                c['overdue_assignee'] = r.assignee_display or '未分配'
        if r.test_cases and r.test_cases > 0:
            c['tc_total'] += r.test_cases
            c['tc_done'] += math.floor(r.test_cases * r.weighted_completion / 100)
        if r.start_date and (not c['min_start'] or r.start_date < c['min_start']):
            c['min_start'] = r.start_date
        if r.due_date and (not c['max_due'] or r.due_date > c['max_due']):
            c['max_due'] = r.due_date
        if r.assignee_display and r.assignee_display != '未分配':
            _src_assignees[key].append(r.assignee_display)

    for key, c in pivot_src_cells.items():
        _finish_cell(c, _src_assignees.get(key, []))

    # Only keep sources that have at least one non-empty cell
    _used_sources = set(r.source or 'coding' for r in pivot_reqs)
    _sources = [s for s in _all_sources if s in _used_sources]

    # Per-L2 active sources: only keep sources that have data under each L2
    _l2_used = defaultdict(set)
    for r in pivot_reqs:
        _l2_used[r.category_l2].add(r.source or 'coding')
    _l2_sources = {l2: [s for s in _sources if s in _l2_used.get(l2, set())] for l2 in pivot_l2s}

    def _merge_cells(cells_list):
        merged = _empty_cell()
        for c in cells_list:
            for k in ('total', 'days', 'done', 'overdue', 'active', 'not_started', 'wpct_sum', 'tc_total', 'tc_done'):
                merged[k] += c[k]
            if c['min_start'] and (not merged['min_start'] or c['min_start'] < merged['min_start']):
                merged['min_start'] = c['min_start']
            if c['max_due'] and (not merged['max_due'] or c['max_due'] > merged['max_due']):
                merged['max_due'] = c['max_due']
            if c['overdue_days'] > merged['overdue_days']:
                merged['overdue_days'] = c['overdue_days']
                merged['overdue_assignee'] = c['overdue_assignee']
        _finish_cell(merged, [])
        return merged

    pivot_cells = {}
    for l1 in pivot_l1s:
        for l2 in pivot_l2s:
            sub = [pivot_src_cells.get((l1, l2, s), _empty_cell()) for s in _sources]
            pivot_cells[(l1, l2)] = _merge_cells(sub)

    pivot_row_totals = {l1: _merge_cells([pivot_cells.get((l1, l2), _empty_cell()) for l2 in pivot_l2s]) for l1 in pivot_l1s}
    pivot_col_totals = {l2: _merge_cells([pivot_cells.get((l1, l2), _empty_cell()) for l1 in pivot_l1s]) for l2 in pivot_l2s}
    pivot_grand = _merge_cells(list(pivot_cells.values()))
    pivot_src_row_totals = {}
    for l1 in pivot_l1s:
        for s in _sources:
            pivot_src_row_totals[(l1, s)] = _merge_cells([pivot_src_cells.get((l1, l2, s), _empty_cell()) for l2 in pivot_l2s])
    pivot_src_col_totals = {}
    for l2 in pivot_l2s:
        for s in _sources:
            pivot_src_col_totals[(l2, s)] = _merge_cells([pivot_src_cells.get((l1, l2, s), _empty_cell()) for l1 in pivot_l1s])
    pivot_src_grand = {}
    for s in _sources:
        pivot_src_grand[s] = _merge_cells([pivot_src_cells.get((l1, l2, s), _empty_cell()) for l1 in pivot_l1s for l2 in pivot_l2s])

    return {
        'pivot_l1s': pivot_l1s, 'pivot_l2s': pivot_l2s,
        'pivot_cells': dict(pivot_cells), 'pivot_src_cells': dict(pivot_src_cells),
        'pivot_row_totals': pivot_row_totals, 'pivot_col_totals': pivot_col_totals,
        'pivot_grand': pivot_grand,
        'pivot_src_row_totals': pivot_src_row_totals, 'pivot_src_col_totals': pivot_src_col_totals,
        'pivot_src_grand': pivot_src_grand,
        'pivot_sources': _sources, 'pivot_l2_sources': _l2_sources, 'pivot_src_labels': _src_labels,
        'pivot_l1_start': _l1_min_start, 'pivot_l2_start': _l2_min_start,
        'pivot_l1_due': _l1_max_due, 'pivot_l2_due': _l2_max_due,
    }


# ---- Stats / Weekly Report / Excel Export ----

@dashboard_bp.route('/stats')
@login_required
def stats():

    cur_tab = request.args.get('tab', 'overview')
    period = request.args.get('period', '1w')  # 1w/2w/1m/2m
    offset = request.args.get('week', 0, type=int)
    cur_project_id = request.args.get('project_id', type=int)
    cur_project_id, cur_project = _guard_hidden_project(cur_project_id)
    # People tab: use period range; overview tab: use current week
    monday, sunday = week_range(offset)
    if cur_tab == 'people':
        period_days = {'1w': 7, '2w': 14, '1m': 30, '2m': 60}.get(period, 7)
        people_start = date.today() - timedelta(days=period_days - 1)
        people_end = date.today()
    else:
        people_start, people_end = monday, sunday

    # Include sub-projects toggle
    has_children = cur_project and cur_project.children
    is_child = cur_project and cur_project.parent_id
    default_sub = '1' if has_children and not is_child else '0'
    include_sub = request.args.get('include_sub', default_sub) == '1'

    sub_project_ids = []
    if include_sub and cur_project and cur_project.children:
        sub_project_ids = [c.id for c in cur_project.children]
    all_pids = [cur_project_id] + sub_project_ids if cur_project_id else []

    data = gather_week_stats(people_start, people_end, project_id=cur_project_id)

    # Requirement stats + deltas
    req_overview_q = Requirement.query.filter(Requirement.parent_id.is_(None))
    if cur_project_id:
        req_overview_q = req_overview_q.filter(Requirement.project_id.in_(all_pids))
    all_reqs = req_overview_q.all()
    weekly_deltas = _compute_weekly_deltas(all_reqs, all_pids, monday, sunday) if all_reqs else {}

    # Weighted completion
    _active = [r for r in all_reqs if r.status != 'closed']
    _comp_pct = lambda r: 100 if r.status == 'done' else (r.completion or 0)
    _comp_w_sum = sum(_comp_pct(r) * (r.estimate_days or 1) for r in _active)
    _comp_d_sum = sum((r.estimate_days or 1) for r in _active)
    completion_weighted = round(_comp_w_sum / _comp_d_sum) if _comp_d_sum else None

    # Risk stats + deltas
    from app.models.risk import Risk
    from datetime import timedelta as _td
    risk_q = Risk.query.filter(Risk.deleted_at.is_(None))
    if cur_project_id:
        risk_q = risk_q.filter(Risk.project_id.in_(all_pids))
    if g.hidden_pids:
        risk_q = risk_q.filter(Risk.project_id.notin_(g.hidden_pids))
    all_risks_stats = risk_q.all()
    today_ = date.today()
    _rn = sum(1 for r in all_risks_stats if r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday)
    _rr = sum(1 for r in all_risks_stats if r.status in ('resolved', 'closed') and r.updated_at and r.updated_at.date() >= monday and r.updated_at.date() <= sunday)
    _ro = sum(1 for r in all_risks_stats if r.status == 'open' and r.due_date and monday <= r.due_date < today_)
    risk_stats = {
        'total': len(all_risks_stats),
        'open': sum(1 for r in all_risks_stats if r.status == 'open'),
        'overdue': sum(1 for r in all_risks_stats if r.is_overdue),
        'resolved': sum(1 for r in all_risks_stats if r.status == 'resolved'),
        'closed': sum(1 for r in all_risks_stats if r.status == 'closed'),
        'high': sum(1 for r in all_risks_stats if r.status == 'open' and r.severity == 'high'),
        'high_total': sum(1 for r in all_risks_stats if r.severity == 'high'),
        'medium': sum(1 for r in all_risks_stats if r.status == 'open' and r.severity == 'medium'),
        'medium_total': sum(1 for r in all_risks_stats if r.severity == 'medium'),
        'low': sum(1 for r in all_risks_stats if r.status == 'open' and r.severity == 'low'),
        'low_total': sum(1 for r in all_risks_stats if r.severity == 'low'),
        'new_delta': _rn, 'resolved_delta': _rr, 'overdue_delta': _ro,
        'high_delta': sum(1 for r in all_risks_stats if r.severity == 'high' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
        'medium_delta': sum(1 for r in all_risks_stats if r.severity == 'medium' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
        'low_delta': sum(1 for r in all_risks_stats if r.severity == 'low' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
    }
    from collections import defaultdict as _dd
    domain_stats = _dd(lambda: {'total': 0, 'open': 0})
    for r in all_risks_stats:
        d = (r.domain_display or '').strip() or '未分类'
        domain_stats[d]['total'] += 1
        if r.status == 'open':
            domain_stats[d]['open'] += 1
    domain_stats = {k: v for k, v in sorted(domain_stats.items(), key=lambda x: (-x[1]['open'], -x[1]['total']))}

    # Trends: 12 weeks snapshot
    import json as _json
    _trend_weeks = 12
    _snap_dates = [today_ - timedelta(days=7 * w) for w in range(_trend_weeks - 1, -1, -1)]
    _snap_labels = [d.strftime('%m-%d') for d in _snap_dates]

    # Requirement trend: total, done, overdue per week
    req_trend = {'weeks': _snap_labels, 'total': [], 'done': [], 'overdue': []}
    if all_reqs:
        for snap in _snap_dates:
            total_at = sum(1 for r in all_reqs if r.created_at and r.created_at.date() <= snap)
            done_at = sum(1 for r in all_reqs
                          if r.created_at and r.created_at.date() <= snap
                          and r.status in ('done', 'closed')
                          and (not r.updated_at or r.updated_at.date() <= snap
                               or r.status in ('done', 'closed')))
            # More accurate done_at: count reqs that were done by snap_date
            done_at2 = 0
            for r in all_reqs:
                if not r.created_at or r.created_at.date() > snap:
                    continue
                if r.status in ('done', 'closed'):
                    if r.updated_at and r.updated_at.date() <= snap:
                        done_at2 += 1
                    elif not r.updated_at:
                        done_at2 += 1
                # else: not done now, so wasn't done at snap either (unless re-opened, ignore)
            overdue_at = sum(1 for r in all_reqs
                             if r.created_at and r.created_at.date() <= snap
                             and r.due_date and r.due_date < snap
                             and not (r.status in ('done', 'closed') and r.updated_at and r.updated_at.date() <= snap))
            req_trend['total'].append(total_at)
            req_trend['done'].append(done_at2)
            req_trend['overdue'].append(overdue_at)

    # Risk trend: open count per domain
    risk_trend = {'weeks': _snap_labels, 'domains': {}}
    if all_risks_stats:
        all_domains = sorted(domain_stats.keys())
        for domain in all_domains:
            risk_trend['domains'][domain] = []
            for snap in _snap_dates:
                open_at = 0
                for r in all_risks_stats:
                    rd = (r.domain_display or '').strip() or '未分类'
                    if rd != domain:
                        continue
                    if not r.created_at or r.created_at.date() > snap:
                        continue
                    if r.status == 'open':
                        open_at += 1
                    elif r.updated_at and r.updated_at.date() > snap:
                        open_at += 1
                risk_trend['domains'][domain].append(open_at)

    pivot = _build_pivot_data(cur_project_id, include_sub=include_sub)

    return render_template('dashboard/stats.html',
        data=data, monday=monday, sunday=sunday,
        offset=offset, cur_tab=cur_tab, period=period,
        people_start=people_start, people_end=people_end,
        cur_project=cur_project, cur_project_id=cur_project_id or 0,
        include_sub=include_sub,
        all_reqs=all_reqs, weekly_deltas=weekly_deltas,
        completion_weighted=completion_weighted,
        risk_stats=risk_stats, domain_stats=domain_stats,
        req_trend_json=_json.dumps(req_trend, ensure_ascii=False),
        risk_trend=risk_trend, risk_trend_json=_json.dumps(risk_trend, ensure_ascii=False),
        **pivot,
    )


@dashboard_bp.route('/metrics')
@login_required
def metrics():
    """Delivery cycle time and estimate-vs-actual deviation analytics."""
    import statistics as _stats

    cur_project_id = request.args.get('project_id', type=int)
    cur_project_id, cur_project = _guard_hidden_project(cur_project_id)
    projects = _visible_projects()

    delivery = get_delivery_metrics(project_id=cur_project_id)
    deviation = get_estimate_deviation(project_id=cur_project_id)

    # Delivery summary
    lead_times = [d['lead_time'] for d in delivery if d['lead_time'] is not None]
    cycle_times = [d['cycle_time'] for d in delivery if d['cycle_time'] is not None]
    avg_lead = round(sum(lead_times) / len(lead_times), 1) if lead_times else 0
    avg_cycle = round(sum(cycle_times) / len(cycle_times), 1) if cycle_times else 0

    # Deviation summary
    dev_pcts = [d['deviation_pct'] for d in deviation]
    avg_deviation = round(sum(dev_pcts) / len(dev_pcts), 1) if dev_pcts else 0
    median_deviation = round(_stats.median(dev_pcts), 1) if dev_pcts else 0

    return render_template('dashboard/metrics.html',
        delivery=delivery, deviation=deviation,
        avg_lead=avg_lead, avg_cycle=avg_cycle,
        avg_deviation=avg_deviation, median_deviation=median_deviation,
        projects=projects, cur_project=cur_project,
        cur_project_id=cur_project_id or 0,
    )


@dashboard_bp.route('/stats/export')
@login_required
def stats_export():
    """Export weekly stats as Excel."""
    import openpyxl
    from openpyxl.styles import Font

    offset = request.args.get('week', 0, type=int)
    cur_group = request.args.get('group', '')
    cur_project_id = request.args.get('project_id', type=int)
    cur_project_id, _ = _guard_hidden_project(cur_project_id)
    monday, sunday = week_range(offset)
    data = gather_week_stats(monday, sunday, group=cur_group or None, project_id=cur_project_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '周报统计'

    # Header
    ws.append([f'周报统计 {monday} ~ {sunday}'])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])

    # Requirement summary
    ws.append(['需求完成率', f'{data.req_done}/{data.req_total}', f'{data.req_rate}%'])
    ws.append([])

    # Per-user table
    ws.append(['姓名', '工号', '新建任务', '完成任务', '完成率'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for s in data.user_stats:
        ws.append([s.user.name, s.user.employee_id or '', s.created, s.done, f'{s.rate}%'])

    # Project investment sheets
    for pi in data.project_investment:
        ws2 = wb.create_sheet(title=pi.project.name[:30])
        ws2.append([f'项目: {pi.project.name}'])
        ws2['A1'].font = Font(bold=True, size=12)
        ws2.append([f'合计: {pi.people_count}人 · {pi.total_days}人天'])
        ws2.append([])
        ws2.append(['姓名', '工号', '投入(人天)', '占比'])
        for cell in ws2[ws2.max_row]:
            cell.font = Font(bold=True)
        for ud in pi.user_days:
            pct = round(ud.days / pi.total_days * 100) if pi.total_days else 0
            ws2.append([ud.user.name, ud.user.employee_id or '', ud.days, f'{pct}%'])
        for col in ws2.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws2.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    # Auto column width for main sheet
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 10)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'weekly_stats_{monday}_{sunday}.xlsx'
    return send_file(buf, download_name=filename,
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _compute_default_recipients(cur_project_id):
    """Compute default To/Cc — delegates to shared utility."""
    from app.utils.recipients import compute_default_recipients
    return compute_default_recipients(cur_project_id)


@dashboard_bp.route('/weekly-report', methods=['GET', 'POST'])
@login_required
def weekly_report():

    offset = request.args.get('week', 0, type=int)
    cur_project_id = request.args.get('project_id', type=int)
    cur_project_id, cur_project = _guard_hidden_project(cur_project_id)
    # Default: include_sub=1 for parent projects, 0 for child projects
    has_children = cur_project and cur_project.children
    is_child = cur_project and cur_project.parent_id
    default_sub = '1' if has_children and not is_child else '0'
    include_sub = request.args.get('include_sub', default_sub) == '1'
    monday, sunday = week_range(offset)

    # Build project ID list (current + children if include_sub)
    sub_project_ids = []
    if include_sub and cur_project and cur_project.children:
        sub_project_ids = [c.id for c in cur_project.children]

    if request.method == 'POST':
        WR_check = WeeklyReport
        frozen = WR_check.query.filter_by(project_id=cur_project_id, week_start=monday, is_frozen=True).first()
        if frozen:
            flash('周报已冻结，无法重新生成', 'warning')
            return redirect(url_for('dashboard.weekly_report', week=offset, project_id=cur_project_id))

        # 1. Completed todos this week
        all_pids = [cur_project_id] + sub_project_ids
        done_q = Todo.query.filter(
            Todo.done_date >= monday, Todo.done_date <= sunday,
        ).options(joinedload(Todo.user), joinedload(Todo.requirements))
        if cur_project_id:
            done_q = done_q.join(todo_requirements, Todo.id == todo_requirements.c.todo_id)\
                           .join(Requirement, Requirement.id == todo_requirements.c.requirement_id)\
                           .filter(Requirement.project_id.in_(all_pids))
        todos_done = done_q.all()

        # 2. Still active todos
        active_q = Todo.query.filter(
            Todo.created_date <= sunday, Todo.status == 'todo',
        ).options(joinedload(Todo.user), joinedload(Todo.requirements))
        if cur_project_id:
            active_q = active_q.join(todo_requirements, Todo.id == todo_requirements.c.todo_id)\
                               .join(Requirement, Requirement.id == todo_requirements.c.requirement_id)\
                               .filter(Requirement.project_id.in_(all_pids))
        todos_active = active_q.all()

        # 3. Requirement changes
        req_q = Requirement.query.filter(
            Requirement.updated_at >= monday,
            Requirement.updated_at <= sunday + timedelta(days=1),
        )
        if cur_project_id:
            req_q = req_q.filter(Requirement.project_id.in_(all_pids))
        req_changes = req_q.all()

        # 4. Per-requirement investment: count distinct people and todo-days
        req_investment = {}  # req_number -> {people: set, days: int}
        for t in todos_done:
            for r in t.requirements:
                inv = req_investment.setdefault(r.number, {'title': r.title, 'people': set(), 'days': 0})
                inv['people'].add(t.user.name)
                inv['days'] += 1
        for t in todos_active:
            for r in t.requirements:
                inv = req_investment.setdefault(r.number, {'title': r.title, 'people': set(), 'days': 0})
                inv['people'].add(t.user.name)
                # Count active days within this week
                start = max(t.created_date, monday)
                end = min(date.today(), sunday)
                inv['days'] += max((end - start).days, 1)

        # 5. All requirements overview (for project scope)
        req_overview_q = Requirement.query.filter(Requirement.parent_id.is_(None))
        if cur_project_id:
            all_pids = [cur_project_id] + sub_project_ids
            req_overview_q = req_overview_q.filter(Requirement.project_id.in_(all_pids))
        all_reqs = _urgency_sort(req_overview_q.all(), limit=50)
        # When include_sub, group by project: parent project first, then children
        if sub_project_ids:
            pid_order = {pid: i for i, pid in enumerate(all_pids)}
            all_reqs.sort(key=lambda r: pid_order.get(r.project_id, 999))

        # 6. Per-person stats this week
        from collections import Counter
        person_done = Counter()
        person_active = Counter()
        for t in todos_done:
            person_done[t.user.name] += 1
        for t in todos_active:
            person_active[t.user.name] += 1

        # 7. Milestones
        milestones = []
        if cur_project:
            milestones = cur_project.milestones

        # Build context
        project_name = cur_project.name if cur_project else '研发团队'
        lines = [f'本周（{monday} ~ {sunday}）{project_name}工作数据：\n']

        # Project goal
        if cur_project and cur_project.description:
            lines.append(f'项目目标：{cur_project.description}\n')

        # Milestones
        if milestones:
            lines.append('里程碑：')
            for ms in milestones:
                due_str = f'，截止 {ms.due_date.strftime("%m-%d")}' if ms.due_date else ''
                status = '已完成' if ms.status == 'completed' else '进行中'
                lines.append(f'- {ms.name}（{status}{due_str}）')

        # Requirement overview
        if all_reqs:
            lines.append('需求总览：')
            for r in all_reqs:
                due_str = f'，预期 {r.due_date.strftime("%m-%d")}' if r.due_date else ''
                days_str = f'，预估 {r.estimate_days}人天' if r.estimate_days else ''
                if r.due_date and r.due_date < date.today() and r.status not in REQ_INACTIVE_STATUSES:
                    overdue = f'⚠️已延期{(date.today() - r.due_date).days}天'
                else:
                    overdue = ''
                children_str = ''
                if r.children:
                    done_children = sum(1 for c in r.children if c.status in REQ_INACTIVE_STATUSES)
                    children_str = f'，子需求 {done_children}/{len(r.children)} 完成'
                assignee = r.assignee_display
                proj_tag = f'[{r.project.name}] ' if sub_project_ids and r.project else ''
                lines.append(f'- {proj_tag}[{r.number}] {r.title}（{r.status_label}，{assignee}{days_str}{due_str}{children_str}）{overdue}')
                for c in (r.children or []):
                    c_due = f'，预期 {c.due_date.strftime("%m-%d")}' if c.due_date else ''
                    c_days = f'，预估 {c.estimate_days}人天' if c.estimate_days else ''
                    c_overdue = ''
                    if c.due_date and c.due_date < date.today() and c.status not in REQ_INACTIVE_STATUSES:
                        c_overdue = f'⚠️已延期{(date.today() - c.due_date).days}天'
                    lines.append(f'  - ↳[{c.number}] {c.title}（{c.status_label}，{c.assignee_display}{c_days}{c_due}）{c_overdue}')

        if todos_done:
            lines.append('\n本周已完成的任务：')
            for t in todos_done:
                reqs_str = ', '.join(r.number for r in t.requirements)
                lines.append(f'- {t.user.name}: {t.title}（{reqs_str}）')

        if todos_active:
            lines.append('\n进行中的任务：')
            for t in todos_active:
                reqs_str = ', '.join(r.number for r in t.requirements)
                lines.append(f'- {t.user.name}: {t.title}（{reqs_str}）')

        # Open risks from DB (needed for both prompt context and template)
        risk_q = Risk.query.filter_by(status='open').filter(Risk.deleted_at.is_(None))
        if cur_project_id:
            all_pids = [cur_project_id] + sub_project_ids
            risk_q = risk_q.filter(Risk.project_id.in_(all_pids))
        open_risks = risk_q.order_by(Risk.severity, Risk.due_date).all()

        # Recently closed risks (last 14 days)
        two_weeks_ago = monday - timedelta(days=14)
        closed_risk_q = Risk.query.filter(Risk.status.in_(['resolved', 'closed']),
            Risk.updated_at >= two_weeks_ago, Risk.deleted_at.is_(None))
        if cur_project_id:
            closed_risk_q = closed_risk_q.filter(Risk.project_id.in_(all_pids))
        recent_closed_risks = closed_risk_q.order_by(Risk.updated_at.desc()).all()

        if open_risks:
            lines.append('\n风险&问题（未解决）：')
            for r in open_risks:
                r_days = (r.due_date - date.today()).days if r.due_date else 999
                r_status = f'已延期{-r_days}天' if r_days < 0 else f'剩{r_days}天'
                owner_info = f'{r.owner}' if r.owner else '无'
                if r.owner_user:
                    owner_info += f' {r.owner_user.employee_id}'
                tracker_info = f'{r.tracker.name} {r.tracker.employee_id}' if r.tracker else (r.tracker_name or '无')
                lines.append(f'- {r.title}（{r.severity_label}，{r_status}，责任人：{owner_info}，跟踪人：{tracker_info}）')

        if req_changes:
            lines.append('\n本周需求状态变更：')
            for r in req_changes:
                lines.append(f'- [{r.number}] {r.title}（{r.status_label}）')

        if req_investment:
            lines.append('\n需求投入汇总：')
            for num, inv in sorted(req_investment.items()):
                people_list = ', '.join(sorted(inv['people']))
                # Use estimate_days from requirement if available
                req_obj = next((r for r in all_reqs if r.number == num), None)
                est = f'预估{req_obj.estimate_days}人天' if req_obj and req_obj.estimate_days else f'{inv["days"]}个任务'
                lines.append(f'- [{num}] {inv["title"]}: {len(inv["people"])}人，{est}（{people_list}）')

        # Person stats
        all_persons = sorted(set(list(person_done.keys()) + list(person_active.keys())))
        if all_persons:
            lines.append('\n人员投入：')
            for name in all_persons:
                lines.append(f'- {name}: 完成 {person_done.get(name, 0)} 个任务，进行中 {person_active.get(name, 0)} 个')

        # Sub-project context for AI
        if cur_project and cur_project.children:
            lines.append('\n专题（子项目）进展：')
            for child in cur_project.children:
                child_reqs = Requirement.query.filter_by(project_id=child.id, parent_id=None).all()
                c_total = len(child_reqs)
                c_done = sum(1 for r in child_reqs if r.status in ('done', 'closed'))
                from app.models.project_member import ProjectMember as PM_
                pm = PM_.query.filter_by(project_id=child.id, project_role='PM').first()
                fo = pm or PM_.query.filter_by(project_id=child.id, project_role='FO').first()
                fo_name = fo.display_name if fo else (child.owner.name if child.owner else '未分配')
                lines.append(f'- {child.name}（负责人：{fo_name}，需求 {c_done}/{c_total} 完成，进度 {child.progress}%）')

        import json as json_lib

        # Check if AI service is enabled
        ai_enabled = current_app.config.get('AI_ENABLED', False)

        if ai_enabled:
            # AI prompt: only generate analysis (summary, risks, plan)
            tpl = get_prompt('weekly_report')
            prompt = tpl.format(project_name=project_name) + '\n\n' + '\n'.join(lines)

            result, raw = call_ollama(prompt)
            if isinstance(result, dict):
                ai_analysis = {
                    'summary': result.get('summary', '数据不足，无法生成摘要'),
                    'highlights': result.get('highlights', []),
                    'risks': result.get('risks', []),
                    'plan': result.get('plan', []),
                }
            elif result is None:
                ai_analysis = {
                    'summary': 'AI服务暂时不可用，请人工填写',
                    'highlights': [],
                    'risks': [],
                    'plan': [],
                }
            else:
                ai_analysis = {
                    'summary': '数据不足，无法生成摘要',
                    'highlights': [],
                    'risks': [],
                    'plan': [],
                }
        else:
            # Rule-based summary when AI is disabled
            ai_analysis = _build_rule_based_summary(
                project_name, all_reqs, todos_done, todos_active,
                open_risks, milestones, monday, sunday,
            )

        # Reviewer: PL of current user's group; if user is PL, then XM; fallback to manager
        reviewer = ''
        if current_user.has_role('PL'):
            xm_users = User.query.filter(User.is_active == True, User.group == current_user.group)\
                .join(User.roles).filter(Role.name.in_(['XM', 'PM'])).first()
            reviewer = xm_users.name if xm_users else ''
        else:
            pl_user = User.query.filter(User.is_active == True, User.group == current_user.group)\
                .join(User.roles).filter(Role.name == 'PL').first()
            reviewer = pl_user.name if pl_user else ''
        if not reviewer and current_user.manager:
            parts = current_user.manager.strip().split()
            reviewer = parts[0] if parts else '待定'
        if not reviewer:
            reviewer = '待定'

        # People map: person × requirement matrix
        people_map = {}
        people_map_reqs = sorted(req_investment.keys())
        for rnum, inv in req_investment.items():
            for pname in inv['people']:
                if pname not in people_map:
                    people_map[pname] = {'_total': 0}
                people_map[pname][rnum] = inv['days'] // max(len(inv['people']), 1)
                people_map[pname]['_total'] += people_map[pname][rnum]

        # Generate milestone timeline image (base64 PNG)
        timeline_img = None
        if milestones:
            try:
                from app.services.timeline import generate_timeline_image
                ms_data = [{'name': m.name, 'due_date': m.due_date, 'status': m.status} for m in milestones]
                timeline_img = generate_timeline_image(ms_data)
            except Exception:  # noqa: S110
                pass

        # Smart requirement list: multi-tier filtering
        # Filter children by project scope; also find orphan children
        all_pids_set = set(all_pids) if cur_project_id else None
        all_with_children = []
        seen_ids = set()
        for r in all_reqs:
            all_with_children.append(r)
            seen_ids.add(r.id)
            for c in (r.children or []):
                if all_pids_set and c.project_id not in all_pids_set:
                    continue  # child in different project, skip
                all_with_children.append(c)
                seen_ids.add(c.id)
        # Orphan children: in this project but parent is in another project
        if cur_project_id:
            orphans = Requirement.query.filter(
                Requirement.parent_id.isnot(None),
                Requirement.project_id.in_(all_pids),
                Requirement.id.notin_(seen_ids) if seen_ids else True,
            ).all()
            all_with_children.extend(orphans)
        # Full mode: include parents + all children
        display_reqs = list(all_with_children)
        req_list_mode = 'full'  # full / diff_assignee / parent_only / priority

        if len(all_with_children) > 40:
            # Tier 2: parent + children with different assignee
            display_reqs = []
            for r in all_reqs:
                display_reqs.append(r)
                for c in (r.children or []):
                    if c.assignee_id != r.assignee_id:
                        display_reqs.append(c)
            req_list_mode = 'diff_assignee'

        if len(display_reqs) > 40:
            # Tier 3: parent only
            display_reqs = list(all_reqs)
            req_list_mode = 'parent_only'

        if len(display_reqs) > 40:
            # Tier 4: urgency sort with limit
            display_reqs = _urgency_sort(display_reqs, limit=40)
            req_list_mode = 'priority'

        # Package all data for template and Excel
        sub_projects = _build_sub_projects(cur_project, monday)

        # Risk stats & domain stats for report
        # Stats scope: open + recently closed (2 weeks)
        all_project_risks = list(open_risks) + list(recent_closed_risks)
        # Deltas: this week's new/resolved/new overdue
        _risk_new = sum(1 for r in all_project_risks if r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday)
        _risk_resolved = sum(1 for r in all_project_risks if r.status in ('resolved', 'closed') and r.updated_at and r.updated_at.date() >= monday and r.updated_at.date() <= sunday)
        _risk_new_overdue = sum(1 for r in all_project_risks if r.status == 'open' and r.due_date and monday <= r.due_date < date.today())
        risk_stats = {
            'total': len(all_project_risks),
            'open': sum(1 for r in all_project_risks if r.status == 'open'),
            'overdue': sum(1 for r in all_project_risks if r.is_overdue),
            'resolved': sum(1 for r in all_project_risks if r.status == 'resolved'),
            'closed': sum(1 for r in all_project_risks if r.status == 'closed'),
            'high': sum(1 for r in all_project_risks if r.status == 'open' and r.severity == 'high'),
            'high_total': sum(1 for r in all_project_risks if r.severity == 'high'),
            'medium': sum(1 for r in all_project_risks if r.status == 'open' and r.severity == 'medium'),
            'medium_total': sum(1 for r in all_project_risks if r.severity == 'medium'),
            'low': sum(1 for r in all_project_risks if r.status == 'open' and r.severity == 'low'),
            'low_total': sum(1 for r in all_project_risks if r.severity == 'low'),
            'new_delta': _risk_new,
            'resolved_delta': _risk_resolved,
            'overdue_delta': _risk_new_overdue,
            'high_delta': sum(1 for r in all_project_risks if r.severity == 'high' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
            'medium_delta': sum(1 for r in all_project_risks if r.severity == 'medium' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
            'low_delta': sum(1 for r in all_project_risks if r.severity == 'low' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
        }
        from collections import defaultdict as _defaultdict
        domain_stats = _defaultdict(lambda: {'total': 0, 'open': 0})
        for r in all_project_risks:
            domain = (r.domain_display or '未分类')
            domain_stats[domain]['total'] += 1
            if r.status == 'open':
                domain_stats[domain]['open'] += 1
        domain_stats = {k: v for k, v in sorted(domain_stats.items(), key=lambda x: (-x[1]['open'], -x[1]['total']))}

        # Weighted completion progress
        _active = [r for r in all_reqs if r.status != 'closed']
        _comp_pct = lambda r: 100 if r.status == 'done' else (r.completion or 0)
        _comp_w_sum = sum(_comp_pct(r) * (r.estimate_days or 1) for r in _active)
        _comp_d_sum = sum((r.estimate_days or 1) for r in _active)
        completion_weighted = round(_comp_w_sum / _comp_d_sum) if _comp_d_sum else None
        weekly_deltas = _compute_weekly_deltas(all_reqs, all_pids, monday, sunday)

        report_data = {
            'project_name': project_name,
            'project_goal': cur_project.description if cur_project else '',
            'today': date.today(),
            'monday': monday,
            'sunday': sunday,
            'reviewer': reviewer,
            'milestones': milestones,
            'all_reqs': all_reqs, 'display_reqs': display_reqs, 'req_list_mode': req_list_mode, 'req_total_with_children': len(all_with_children),
            'req_investment': req_investment,
            'completion_weighted': completion_weighted,
            'weekly_deltas': weekly_deltas,
            'person_done': dict(person_done),
            'person_active': dict(person_active),
            'all_persons': all_persons,
            'todos_done': todos_done,
            'todos_active': todos_active,
            'open_risks': open_risks,
            'recent_closed_risks': recent_closed_risks,
            'risk_stats': risk_stats,
            'domain_stats': domain_stats,
            'people_map': people_map,
            'people_map_reqs': people_map_reqs,
            'people_roles': _merge_member_roles(ProjectMember.query.filter(ProjectMember.project_id.in_([cur_project_id] + sub_project_ids)).all()),
            'people_tree': _build_people_tree(cur_project_id, sub_project_ids),
            'people_tree_img': _gen_people_tree_img(cur_project_id, sub_project_ids, project_name),
            'ai': ai_analysis,
            'sub_projects': sub_projects,
            'timeline_img': timeline_img,
        }

        # Save to DB
        saved = WeeklyReport.query.filter_by(project_id=cur_project_id, week_start=monday).first()
        if saved:
            saved.summary = ai_analysis['summary']
            saved.risks_json = json_lib.dumps(ai_analysis['risks'], ensure_ascii=False)
            saved.plan_json = json_lib.dumps(ai_analysis['plan'], ensure_ascii=False)
            saved.updated_at = datetime.now()
        else:
            saved = WeeklyReport(
                project_id=cur_project_id,
                week_start=monday, week_end=sunday,
                summary=ai_analysis['summary'],
                risks_json=json_lib.dumps(ai_analysis['risks'], ensure_ascii=False),
                plan_json=json_lib.dumps(ai_analysis['plan'], ensure_ascii=False),
                created_by=current_user.id,
            )
            db.session.add(saved)
        db.session.commit()

        _def_to, _def_cc = _compute_default_recipients(cur_project_id)
        pivot = _build_pivot_data(cur_project_id, include_sub=include_sub)
        return render_template('dashboard/weekly_report.html',
            report_data=report_data, saved_report=saved,
            monday=monday, sunday=sunday, offset=offset,
            cur_project=cur_project, cur_project_id=cur_project_id or 0,
            include_sub=include_sub,
            default_to=_def_to, default_cc=_def_cc,
            **pivot,
            pinned_knowledge=_get_pinned_knowledge(cur_project_id, sub_project_ids if include_sub else []),
        )

    # GET: check if saved report exists, and load full DB data
    import json as json_lib
    saved = WeeklyReport.query.filter_by(project_id=cur_project_id, week_start=monday).first() if cur_project_id else None

    if saved:
        # Load saved AI analysis + fresh DB data
        from collections import Counter

        project_name = cur_project.name if cur_project else '研发团队'

        # Milestones
        milestones = cur_project.milestones if cur_project else []

        # Requirements
        req_overview_q = Requirement.query.filter(Requirement.parent_id.is_(None))
        if cur_project_id:
            all_pids = [cur_project_id] + sub_project_ids
            req_overview_q = req_overview_q.filter(Requirement.project_id.in_(all_pids))
        all_pids_set = set(all_pids) if cur_project_id else None
        all_reqs = _urgency_sort(req_overview_q.all(), limit=50)
        if sub_project_ids:
            pid_order = {pid: i for i, pid in enumerate(all_pids)}
            all_reqs.sort(key=lambda r: pid_order.get(r.project_id, 999))

        # Todos
        done_q = Todo.query.filter(Todo.done_date >= monday, Todo.done_date <= sunday)\
            .options(joinedload(Todo.user), joinedload(Todo.requirements))
        active_q = Todo.query.filter(Todo.created_date <= sunday, Todo.status == 'todo')\
            .options(joinedload(Todo.user), joinedload(Todo.requirements))
        if cur_project_id:
            done_q = done_q.join(todo_requirements, Todo.id == todo_requirements.c.todo_id)\
                           .join(Requirement, Requirement.id == todo_requirements.c.requirement_id)\
                           .filter(Requirement.project_id == cur_project_id)
            active_q = active_q.join(todo_requirements, Todo.id == todo_requirements.c.todo_id)\
                               .join(Requirement, Requirement.id == todo_requirements.c.requirement_id)\
                               .filter(Requirement.project_id == cur_project_id)
        todos_done = done_q.all()
        todos_active = active_q.all()

        person_done = Counter(t.user.name for t in todos_done)
        person_active = Counter(t.user.name for t in todos_active)
        all_persons = sorted(set(list(person_done.keys()) + list(person_active.keys())))

        req_investment = {}
        for t in todos_done + todos_active:
            for r in t.requirements:
                inv = req_investment.setdefault(r.number, {'title': r.title, 'people': set(), 'days': 0})
                inv['people'].add(t.user.name)
                inv['days'] += 1

        risk_q = Risk.query.filter_by(status='open').filter(Risk.deleted_at.is_(None))
        if cur_project_id:
            all_pids = [cur_project_id] + sub_project_ids
            risk_q = risk_q.filter(Risk.project_id.in_(all_pids))
        open_risks = risk_q.order_by(Risk.severity, Risk.due_date).all()

        # Recently closed risks (last 14 days)
        two_weeks_ago = monday - timedelta(days=14)
        closed_risk_q = Risk.query.filter(Risk.status.in_(['resolved', 'closed']),
            Risk.updated_at >= two_weeks_ago, Risk.deleted_at.is_(None))
        if cur_project_id:
            closed_risk_q = closed_risk_q.filter(Risk.project_id.in_(all_pids))
        recent_closed_risks = closed_risk_q.order_by(Risk.updated_at.desc()).all()

        # Reviewer: PL of current user's group; if user is PL, then XM; fallback to manager
        reviewer = ''
        if current_user.has_role('PL'):
            xm = User.query.filter(User.is_active == True, User.group == current_user.group)\
                .join(User.roles).filter(Role.name.in_(['XM', 'PM'])).first()
            reviewer = xm.name if xm else ''
        else:
            pl = User.query.filter(User.is_active == True, User.group == current_user.group)\
                .join(User.roles).filter(Role.name == 'PL').first()
            reviewer = pl.name if pl else ''
        if not reviewer and current_user.manager:
            parts = current_user.manager.strip().split()
            reviewer = parts[0] if parts else '待定'
        if not reviewer:
            reviewer = '待定'

        # People map
        people_map = {}
        people_map_reqs = sorted(req_investment.keys())
        for rnum, inv in req_investment.items():
            for pname in inv['people']:
                if pname not in people_map:
                    people_map[pname] = {'_total': 0}
                people_map[pname][rnum] = inv['days'] // max(len(inv['people']), 1)
                people_map[pname]['_total'] += people_map[pname][rnum]

        # Merge saved AI analysis
        ai_analysis = {
            'summary': saved.summary or '',
            'risks': json_lib.loads(saved.risks_json) if saved.risks_json else [],
            'plan': json_lib.loads(saved.plan_json) if saved.plan_json else [],
        }

        # Sub-projects
        sub_projects = []
        if cur_project and cur_project.children:
            for child in cur_project.children:
                child_saved = WeeklyReport.query.filter_by(
                    project_id=child.id, week_start=monday).first()
                from app.models.project_member import ProjectMember as PM_
                pm = PM_.query.filter_by(project_id=child.id, project_role='PM').first()
                fo = pm or PM_.query.filter_by(project_id=child.id, project_role='FO').first()
                child_summary = (child_saved.summary if child_saved and child_saved.summary else '') or ''
                sub_projects.append({
                    'project': child,
                    'owner': fo.user if fo and fo.user else (child.owner if child.owner else None),
                    'summary': child_summary or '-',
                })

        # Risk stats & domain stats for saved report (open + recently closed)
        all_project_risks2 = list(open_risks) + list(recent_closed_risks)
        _rn2 = sum(1 for r in all_project_risks2 if r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday)
        _rr2 = sum(1 for r in all_project_risks2 if r.status in ('resolved', 'closed') and r.updated_at and r.updated_at.date() >= monday and r.updated_at.date() <= sunday)
        _ro2 = sum(1 for r in all_project_risks2 if r.status == 'open' and r.due_date and monday <= r.due_date < date.today())
        risk_stats2 = {
            'total': len(all_project_risks2),
            'open': sum(1 for r in all_project_risks2 if r.status == 'open'),
            'overdue': sum(1 for r in all_project_risks2 if r.is_overdue),
            'resolved': sum(1 for r in all_project_risks2 if r.status == 'resolved'),
            'closed': sum(1 for r in all_project_risks2 if r.status == 'closed'),
            'high': sum(1 for r in all_project_risks2 if r.status == 'open' and r.severity == 'high'),
            'high_total': sum(1 for r in all_project_risks2 if r.severity == 'high'),
            'medium': sum(1 for r in all_project_risks2 if r.status == 'open' and r.severity == 'medium'),
            'medium_total': sum(1 for r in all_project_risks2 if r.severity == 'medium'),
            'low': sum(1 for r in all_project_risks2 if r.status == 'open' and r.severity == 'low'),
            'low_total': sum(1 for r in all_project_risks2 if r.severity == 'low'),
            'new_delta': _rn2, 'resolved_delta': _rr2, 'overdue_delta': _ro2,
            'high_delta': sum(1 for r in all_project_risks2 if r.severity == 'high' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
            'medium_delta': sum(1 for r in all_project_risks2 if r.severity == 'medium' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
            'low_delta': sum(1 for r in all_project_risks2 if r.severity == 'low' and r.created_at and r.created_at.date() >= monday and r.created_at.date() <= sunday),
        }
        from collections import defaultdict as _defaultdict
        domain_stats2 = _defaultdict(lambda: {'total': 0, 'open': 0})
        for r in all_project_risks2:
            domain = (r.domain_display or '未分类')
            domain_stats2[domain]['total'] += 1
            if r.status == 'open':
                domain_stats2[domain]['open'] += 1
        domain_stats2 = {k: v for k, v in sorted(domain_stats2.items(), key=lambda x: (-x[1]['open'], -x[1]['total']))}

        # Weighted completion progress
        _active2 = [r for r in all_reqs if r.status != 'closed']
        _comp_pct2 = lambda r: 100 if r.status == 'done' else (r.completion or 0)
        _comp_w_sum2 = sum(_comp_pct2(r) * (r.estimate_days or 1) for r in _active2)
        _comp_d_sum2 = sum((r.estimate_days or 1) for r in _active2)
        completion_weighted2 = round(_comp_w_sum2 / _comp_d_sum2) if _comp_d_sum2 else None
        weekly_deltas2 = _compute_weekly_deltas(all_reqs, all_pids, monday, sunday)

        report_data = {
            'project_name': project_name,
            'project_goal': cur_project.description if cur_project else '',
            'today': date.today(),
            'monday': monday,
            'sunday': sunday,
            'reviewer': reviewer,
            'milestones': milestones,
            'all_reqs': all_reqs,
            'display_reqs': [item for r in all_reqs for item in [r] + [c for c in (r.children or []) if not all_pids_set or c.project_id in all_pids_set]],
            'req_list_mode': 'full',
            'req_total_with_children': sum(1 + sum(1 for c in (r.children or []) if not all_pids_set or c.project_id in all_pids_set) for r in all_reqs),
            'req_investment': req_investment,
            'completion_weighted': completion_weighted2,
            'weekly_deltas': weekly_deltas2,
            'person_done': dict(person_done),
            'person_active': dict(person_active),
            'all_persons': all_persons,
            'todos_done': todos_done,
            'todos_active': todos_active,
            'open_risks': open_risks,
            'recent_closed_risks': recent_closed_risks,
            'risk_stats': risk_stats2,
            'domain_stats': domain_stats2,
            'people_map': people_map,
            'people_map_reqs': people_map_reqs,
            'people_roles': _merge_member_roles(ProjectMember.query.filter(ProjectMember.project_id.in_([cur_project_id] + sub_project_ids)).all()),
            'people_tree': _build_people_tree(cur_project_id, sub_project_ids),
            'people_tree_img': _gen_people_tree_img(cur_project_id, sub_project_ids, project_name),
            'ai': ai_analysis,
            'sub_projects': sub_projects,
        }

        # Generate timeline image for saved report too
        timeline_img_saved = None
        if milestones:
            try:
                from app.services.timeline import generate_timeline_image
                ms_data = [{'name': m.name, 'due_date': m.due_date, 'status': m.status} for m in milestones]
                timeline_img_saved = generate_timeline_image(ms_data)
            except Exception:  # noqa: S110
                pass
        report_data['timeline_img'] = timeline_img_saved

        _def_to, _def_cc = _compute_default_recipients(cur_project_id)
        pivot = _build_pivot_data(cur_project_id, include_sub=include_sub)
        return render_template('dashboard/weekly_report.html',
            report_data=report_data, saved_report=saved,
            monday=monday, sunday=sunday, offset=offset,
            cur_project=cur_project, cur_project_id=cur_project_id or 0,
            include_sub=include_sub,
            default_to=_def_to, default_cc=_def_cc,
            **pivot,
            pinned_knowledge=_get_pinned_knowledge(cur_project_id, sub_project_ids if include_sub else []),
        )

    _def_to, _def_cc = _compute_default_recipients(cur_project_id)
    pivot = _build_pivot_data(cur_project_id, include_sub=include_sub)
    return render_template('dashboard/weekly_report.html',
        report_data=None, saved_report=None,
        monday=monday, sunday=sunday, offset=offset,
        cur_project=cur_project, cur_project_id=cur_project_id or 0,
        include_sub=include_sub,
        default_to=_def_to, default_cc=_def_cc,
        **pivot,
        pinned_knowledge=_get_pinned_knowledge(cur_project_id, sub_project_ids if include_sub else []),
    )


@dashboard_bp.route('/weekly-report/save', methods=['POST'])
@login_required
def weekly_report_save():
    """Save manually edited report content."""
    import json as json_lib

    cur_project_id = request.form.get('project_id', type=int)
    week_start_str = request.form.get('week_start', '')
    if not cur_project_id or not week_start_str:
        flash('参数缺失', 'danger')
        return redirect(request.referrer or url_for('dashboard.weekly_report'))
    week_start = date.fromisoformat(week_start_str)

    saved = WeeklyReport.query.filter_by(project_id=cur_project_id, week_start=week_start).first()
    if not saved:
        flash('请先生成周报', 'warning')
        return redirect(request.referrer or url_for('dashboard.weekly_report'))

    if saved.is_frozen:
        flash('周报已冻结，无法编辑', 'warning')
        offset = request.form.get('offset', 0, type=int)
        return redirect(url_for('dashboard.weekly_report', week=offset, project_id=cur_project_id))

    saved.summary = request.form.get('summary', '').strip()
    risks = [r.strip() for r in request.form.get('risks', '').strip().splitlines() if r.strip()]
    plan = [p.strip() for p in request.form.get('plan', '').strip().splitlines() if p.strip()]
    saved.risks_json = json_lib.dumps(risks, ensure_ascii=False)
    saved.plan_json = json_lib.dumps(plan, ensure_ascii=False)
    saved.updated_at = datetime.now()
    db.session.commit()
    flash('周报已保存', 'success')

    offset = request.form.get('offset', 0, type=int)
    return redirect(url_for('dashboard.weekly_report', week=offset, project_id=cur_project_id))


@dashboard_bp.route('/weekly-report/freeze', methods=['POST'])
@login_required
def weekly_report_freeze():
    """Freeze/unfreeze weekly report. Only project PM (owner) can freeze."""

    cur_project_id = request.form.get('project_id', type=int)
    week_start_str = request.form.get('week_start', '')
    action = request.form.get('action', 'freeze')
    week_start = date.fromisoformat(week_start_str) if week_start_str else None

    saved = WeeklyReport.query.filter_by(project_id=cur_project_id, week_start=week_start).first()
    if not saved:
        flash('周报不存在', 'danger')
        return redirect(request.referrer or url_for('dashboard.weekly_report'))

    project = db.session.get(Project, cur_project_id)
    is_pm = project and project.owner_id == current_user.id
    if not is_pm and not current_user.is_admin:
        flash('只有项目 PM 或管理员可以冻结/解冻周报', 'danger')
        return redirect(request.referrer or url_for('dashboard.weekly_report'))

    if action == 'freeze':
        saved.is_frozen = True
        saved.frozen_by = current_user.id
        saved.frozen_at = datetime.now()
        flash('周报已冻结', 'success')
    else:
        saved.is_frozen = False
        saved.frozen_by = None
        saved.frozen_at = None
        flash('周报已解冻', 'success')

    db.session.commit()
    offset = request.form.get('offset', 0, type=int)
    return redirect(url_for('dashboard.weekly_report', week=offset, project_id=cur_project_id))


@dashboard_bp.route('/weekly-report/export', methods=['POST'])
@login_required
def weekly_report_export():
    """Export weekly report as formatted Excel."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    offset = request.args.get('week', 0, type=int)
    cur_project_id = request.args.get('project_id', type=int)
    cur_project_id, cur_project = _guard_hidden_project(cur_project_id)
    monday, sunday = week_range(offset)
    project_name = cur_project.name if cur_project else '研发团队'

    # AI analysis from form hidden fields
    ai_summary = request.form.get('ai_summary', '')
    ai_risks = [r for r in request.form.get('ai_risks', '').split('||') if r]
    ai_plan = [p for p in request.form.get('ai_plan', '').split('||') if p]

    # Gather data
    req_q = Requirement.query.filter(Requirement.parent_id.is_(None))
    if cur_project_id:
        req_q = req_q.filter_by(project_id=cur_project_id)
    all_reqs = _urgency_sort(req_q.all(), limit=50)

    milestones = cur_project.milestones if cur_project else []

    # Todo stats
    from collections import Counter
    done_q = Todo.query.filter(Todo.done_date >= monday, Todo.done_date <= sunday)
    active_q = Todo.query.filter(Todo.created_date <= sunday, Todo.status == 'todo')
    if cur_project_id:
        done_q = done_q.join(todo_requirements, Todo.id == todo_requirements.c.todo_id)\
                       .join(Requirement, Requirement.id == todo_requirements.c.requirement_id)\
                       .filter(Requirement.project_id == cur_project_id)
        active_q = active_q.join(todo_requirements, Todo.id == todo_requirements.c.todo_id)\
                           .join(Requirement, Requirement.id == todo_requirements.c.requirement_id)\
                           .filter(Requirement.project_id == cur_project_id)
    todos_done = done_q.options(joinedload(Todo.user), joinedload(Todo.requirements)).all()
    todos_active = active_q.options(joinedload(Todo.user), joinedload(Todo.requirements)).all()

    person_done = Counter(t.user.name for t in todos_done)
    person_active = Counter(t.user.name for t in todos_active)
    all_persons = sorted(set(list(person_done.keys()) + list(person_active.keys())))

    req_investment = {}
    for t in todos_done + todos_active:
        for r in t.requirements:
            inv = req_investment.setdefault(r.number, {'title': r.title, 'people': set(), 'days': 0})
            inv['people'].add(t.user.name)
            inv['days'] += 1

    # ---- Build Excel ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '项目周报'

    # Styles
    header_font = Font(bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='4A6CF7', end_color='4A6CF7', fill_type='solid')
    section_font = Font(bold=True, size=11, color='2D3748')
    section_fill = PatternFill(start_color='EDF2F7', end_color='EDF2F7', fill_type='solid')
    th_font = Font(bold=True, size=9)
    th_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    td_font = Font(size=9)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E0'),
        right=Side(style='thin', color='CBD5E0'),
        top=Side(style='thin', color='CBD5E0'),
        bottom=Side(style='thin', color='CBD5E0'),
    )
    risk_font = Font(size=9, color='E53E3E')

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f'{project_name} 项目周报')
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 1

    # Date range
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f'{monday.strftime("%Y-%m-%d")} ~ {sunday.strftime("%Y-%m-%d")}')
    cell.font = Font(size=10, color='718096')
    cell.alignment = Alignment(horizontal='center')
    row += 2

    def write_section(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        row += 1

    def write_table_header(headers):
        nonlocal row
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=h)
            cell.font = th_font
            cell.fill = th_fill
            cell.border = thin_border
        row += 1

    def write_table_row(values):
        nonlocal row
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = td_font
            cell.border = thin_border
        row += 1

    # Summary
    write_section('整体进展')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value=ai_summary).font = Font(size=10)
    row += 2

    # Milestones
    if milestones:
        write_section('里程碑')
        write_table_header(['里程碑', '状态', '截止日期'])
        for ms in milestones:
            write_table_row([
                ms.name,
                '已完成' if ms.status == 'completed' else '进行中',
                ms.due_date.strftime('%Y-%m-%d') if ms.due_date else '-',
            ])
        row += 1

    # Requirements
    if all_reqs:
        write_section('需求进展')
        write_table_header(['标题', '状态', '负责人', '预估(人天)', '预期开始', '预期完成', '本周投入'])
        for r in all_reqs:
            inv = req_investment.get(r.number)
            invest_str = f'{len(inv["people"])}人·{inv["days"]}天' if inv else '-'
            children_str = ''
            if r.children:
                dc = sum(1 for c in r.children if c.status in REQ_INACTIVE_STATUSES)
                children_str = f' ({dc}/{len(r.children)})'
            overdue = ' [超期]' if (r.due_date and r.due_date < date.today() and r.status not in REQ_INACTIVE_STATUSES) else ''
            write_table_row([
                r.title + children_str,
                r.status_label + overdue,
                r.assignee_display,
                r.estimate_days or '-',
                r.start_date.strftime('%m-%d') if r.start_date else '-',
                r.due_date.strftime('%m-%d') if r.due_date else '-',
                invest_str,
            ])
        row += 1

    # Person stats
    if all_persons:
        write_section('人员投入')
        write_table_header(['姓名', '完成任务', '进行中'])
        for name in all_persons:
            write_table_row([name, person_done.get(name, 0), person_active.get(name, 0)])
        row += 1

    # Risks
    write_section('风险与问题')
    if ai_risks:
        for r in ai_risks:
            cell = ws.cell(row=row, column=1, value=f'  · {r}')
            cell.font = risk_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1
    else:
        ws.cell(row=row, column=1, value='  暂无').font = Font(size=9, color='A0AEC0')
        row += 1
    row += 1

    # Plan
    write_section('下周计划')
    if ai_plan:
        for p in ai_plan:
            ws.cell(row=row, column=1, value=f'  · {p}').font = td_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1
    else:
        ws.cell(row=row, column=1, value='  暂无').font = Font(size=9, color='A0AEC0')

    # Column widths
    col_widths = [10, 30, 10, 10, 12, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'{project_name}_周报_{monday}_{sunday}.xlsx'
    return send_file(buf, download_name=filename,
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ---- My Day (我的一天) ----

@dashboard_bp.route('/my-day')
@login_required
def my_day():
    """Personal daily schedule — 5-day Outlook-style timeline based on pomodoro sessions."""
    from app.models.todo import PomodoroSession

    today = date.today()
    # 5 days: today + 4 previous days (today first)
    days = [today - timedelta(days=i) for i in range(5)]

    # Load all pomodoro sessions for the 5-day range
    range_start = datetime(days[-1].year, days[-1].month, days[-1].day)
    range_end = datetime(today.year, today.month, today.day) + timedelta(days=1)
    all_sessions = PomodoroSession.query.join(Todo).filter(
        Todo.user_id == current_user.id,
        PomodoroSession.created_at >= range_start,
        PomodoroSession.created_at < range_end,
    ).options(joinedload(PomodoroSession.todo)).order_by(PomodoroSession.started_at.asc().nullslast()).all()

    # Group sessions by date
    day_data = {}
    for d in days:
        day_data[d] = {
            'sessions': [],
            'timeline': [],
            'total_min': 0,
            'count': 0,
            'done': [],
        }

    for s in all_sessions:
        s_date = (s.started_at or s.created_at).date() if (s.started_at or s.created_at) else None
        if s_date and s_date in day_data:
            dd = day_data[s_date]
            dd['sessions'].append(s)
            dd['total_min'] += s.minutes
            if s.completed:
                dd['count'] += 1
            start = s.started_at or s.created_at
            dd['timeline'].append({
                'start_hour': start.hour,
                'start_min': start.minute,
                'duration': s.minutes,
                'title': s.todo.title if s.todo else '',
                'completed': s.completed,
            })

    # Load done todos per day
    done_todos = Todo.query.filter(
        Todo.user_id == current_user.id,
        Todo.status == 'done',
        Todo.done_date.in_(days),
    ).all()
    for t in done_todos:
        if t.done_date in day_data:
            day_data[t.done_date]['done'].append(t)

    # Aggregate stats
    total_focus = sum(dd['total_min'] for dd in day_data.values())
    total_sessions = sum(dd['count'] for dd in day_data.values())
    total_done = sum(len(dd['done']) for dd in day_data.values())

    # Focus ranking across 5 days
    focus_by_todo = {}
    for s in all_sessions:
        focus_by_todo.setdefault(s.todo_id, {'title': s.todo.title if s.todo else '', 'minutes': 0, 'count': 0})
        focus_by_todo[s.todo_id]['minutes'] += s.minutes
        focus_by_todo[s.todo_id]['count'] += 1
    focus_ranking = sorted(focus_by_todo.values(), key=lambda x: -x['minutes'])[:10]

    weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']

    # ---- Merged from profile_stats (个人效能) ----
    from app.models.incentive import Incentive
    uid = current_user.id
    six_months_ago = today - timedelta(days=180)

    monthly_todos = db.session.query(
        db.func.strftime('%Y-%m', Todo.done_date).label('month'),
        db.func.count(Todo.id),
    ).filter(
        Todo.user_id == uid, Todo.status == TODO_STATUS_DONE,
        Todo.done_date >= six_months_ago,
    ).group_by('month').order_by('month').all()

    monthly_focus = db.session.query(
        db.func.strftime('%Y-%m', Todo.done_date).label('month'),
        db.func.sum(Todo.actual_minutes),
    ).filter(
        Todo.user_id == uid, Todo.status == TODO_STATUS_DONE,
        Todo.done_date >= six_months_ago, Todo.actual_minutes > 0,
    ).group_by('month').order_by('month').all()

    req_count = Requirement.query.filter_by(assignee_id=uid).count()
    req_done = Requirement.query.filter_by(assignee_id=uid, status='done').count()
    incentive_count = db.session.query(db.func.count(Incentive.id)).filter(
        Incentive.status == 'approved', Incentive.nominees.any(id=uid),
    ).scalar() or 0

    year_ago = today - timedelta(days=365)
    heatmap_rows = db.session.query(
        Todo.done_date, db.func.count(Todo.id),
    ).filter(
        Todo.user_id == uid, Todo.status == TODO_STATUS_DONE,
        Todo.done_date >= year_ago,
    ).group_by(Todo.done_date).all()
    heatmap = {str(row[0]): row[1] for row in heatmap_rows}

    all_focus_hours = round((db.session.query(
        db.func.sum(Todo.actual_minutes),
    ).filter(Todo.user_id == uid, Todo.actual_minutes > 0).scalar() or 0) / 60, 1)

    # Activity timer records (meeting/review/break/other)
    from app.models.activity_timer import ActivityTimer
    act_records = ActivityTimer.query.filter(
        ActivityTimer.user_id == current_user.id,
        ActivityTimer.date.in_(days),
    ).order_by(ActivityTimer.started_at).all()
    act_events = []
    act_colors = {'meeting': '#6f42c1', 'review': '#0d6efd', 'writing': '#d63384', 'break': '#fd7e14', 'manual': '#0d6efd', 'other': '#6c757d'}
    act_total_min = sum(a.minutes for a in act_records if a.date == today)
    for a in act_records:
        dt = a.started_at
        act_events.append({
            'date': a.date.strftime('%Y-%m-%d'),
            'start_hour': dt.hour,
            'start_min': dt.minute,
            'duration': a.minutes,
            'title': a.label,
            'color': act_colors.get(a.activity, '#6c757d'),
        })

    return render_template('dashboard/my_day.html',
                           today=today, days=days, day_data=day_data,
                           total_focus=total_focus, total_sessions=total_sessions,
                           total_done=total_done, focus_ranking=focus_ranking,
                           weekday_names=weekday_names,
                           # 个人效能
                           monthly_todos=monthly_todos, monthly_focus=monthly_focus,
                           req_count=req_count, req_done=req_done,
                           incentive_count=incentive_count,
                           heatmap=heatmap, heatmap_start=year_ago,
                           all_focus_hours=all_focus_hours,
                           timedelta=timedelta,
                           cal_events=session.get('_cal_events', []),
                           act_events=act_events, act_total_min=act_total_min)


@dashboard_bp.route('/my-day/import-ics', methods=['POST'])
@login_required
def import_ics():
    """Import Outlook calendar .ics file and store events in session for timeline display."""
    from icalendar import Calendar

    f = request.files.get('ics_file')
    if not f or not f.filename.endswith('.ics'):
        flash('请上传 .ics 文件', 'warning')
        return redirect(url_for('dashboard.my_day'))

    try:
        cal = Calendar.from_ical(f.read())
    except Exception:
        flash('ICS 文件解析失败', 'danger')
        return redirect(url_for('dashboard.my_day'))

    today = date.today()
    events = []
    for component in cal.walk():
        if component.name != 'VEVENT':
            continue
        dtstart = component.get('dtstart')
        dtend = component.get('dtend')
        summary = str(component.get('summary', ''))
        if not dtstart:
            continue
        dt = dtstart.dt
        # Handle all-day events (date, not datetime)
        if isinstance(dt, date) and not isinstance(dt, datetime):
            continue  # skip all-day events for timeline
        # Convert to local naive datetime
        import zoneinfo
        try:
            local_tz = zoneinfo.ZoneInfo('Asia/Shanghai')
        except Exception:
            local_tz = timezone(timedelta(hours=8))
        if dt.tzinfo:
            dt = dt.astimezone(local_tz).replace(tzinfo=None)
        # Duration
        duration_min = 30  # default
        if dtend:
            dte = dtend.dt
            if isinstance(dte, datetime):
                if dte.tzinfo:
                    dte = dte.astimezone(local_tz).replace(tzinfo=None)
                duration_min = max(5, int((dte - dt).total_seconds() / 60))

        events.append({
            'date': dt.strftime('%Y-%m-%d'),
            'start_hour': dt.hour,
            'start_min': dt.minute,
            'duration': duration_min,
            'title': summary[:50],
        })

    # Store in session (keep events within 5-day range)
    day_range = {(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(5)}
    events = [e for e in events if e['date'] in day_range]
    session['_cal_events'] = events
    flash(f'已导入 {len(events)} 个日历事件', 'success')
    return redirect(url_for('dashboard.my_day'))


@dashboard_bp.route('/my-day/sync-exchange', methods=['POST'])
@login_required
def sync_exchange():
    """Pull calendar events from Exchange using credentials passed from client (not stored)."""
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify(ok=False, error='请输入账号和密码')

    from app.models.site_setting import SiteSetting
    exc_cfg = current_app.config.get('EXCHANGE_CONFIG', {})
    server = SiteSetting.get('exchange_server', exc_cfg.get('server', ''))
    domain = SiteSetting.get('exchange_domain', exc_cfg.get('domain', ''))
    if not server:
        return jsonify(ok=False, error='Exchange 服务器未配置，请联系管理员在后台设置')

    try:
        from exchangelib import Account, Configuration, Credentials, DELEGATE, EWSDateTime, EWSTimeZone

        tz = EWSTimeZone('Asia/Shanghai')
        today = date.today()
        start = today - timedelta(days=4)
        end = today + timedelta(days=1)

        creds = Credentials(username=f'{domain}\\{username}' if domain else username, password=password)
        email = data.get('email', '').strip()
        if not email:
            return jsonify(ok=False, error='请先在个人设置中填写邮箱')
        config = Configuration(server=server, credentials=creds)
        account = Account(email, config=config, autodiscover=False, access_type=DELEGATE)

        events = []
        cal_start = tz.localize(EWSDateTime(start.year, start.month, start.day, 0, 0))
        cal_end = tz.localize(EWSDateTime(end.year, end.month, end.day, 23, 59))

        for item in account.calendar.view(start=cal_start, end=cal_end):
            if item.is_all_day:
                continue
            dt = item.start
            if hasattr(dt, 'astimezone'):
                dt = dt.astimezone(tz)
            duration_min = 30
            if item.end:
                dte = item.end
                if hasattr(dte, 'astimezone'):
                    dte = dte.astimezone(tz)
                duration_min = max(5, int((dte - dt).total_seconds() / 60))
            events.append({
                'date': dt.strftime('%Y-%m-%d'),
                'start_hour': dt.hour,
                'start_min': dt.minute,
                'duration': duration_min,
                'title': (item.subject or '')[:50],
            })

        session['_cal_events'] = events
        return jsonify(ok=True, count=len(events))
    except Exception as e:
        msg = str(e)
        if 'Unauthorized' in msg or '401' in msg:
            return jsonify(ok=False, error='账号或密码错误')
        return jsonify(ok=False, error=f'连接失败: {msg[:100]}')


@dashboard_bp.route('/my-day/clear-ics', methods=['POST'])
@login_required
def clear_ics():
    """Clear imported calendar events from session."""
    session.pop('_cal_events', None)
    flash('已清除导入的日历事件', 'info')
    return redirect(url_for('dashboard.my_day'))


# ---- Personal weekly report ----

@dashboard_bp.route('/my-weekly', methods=['GET', 'POST'])
@login_required
def my_weekly():
    offset = request.args.get('week', 0, type=int)
    monday, sunday = week_range(offset)

    # Completed todos: prefer done_date, fall back to created_date for those without done_date
    my_done = Todo.query.filter_by(user_id=current_user.id, status='done')\
        .filter(db.or_(
            db.and_(Todo.done_date >= monday, Todo.done_date <= sunday),
            db.and_(Todo.done_date.is_(None), Todo.created_date >= monday, Todo.created_date <= sunday),
        ))\
        .options(joinedload(Todo.requirements), joinedload(Todo.items)).all()
    my_active = Todo.query.filter_by(user_id=current_user.id)\
        .filter(Todo.status == 'todo', Todo.created_date <= sunday)\
        .options(joinedload(Todo.requirements), joinedload(Todo.items)).all()

    my_reqs = set()
    for t in my_done + my_active:
        for r in t.requirements:
            my_reqs.add(r)
    # Also include requirements (and child requirements) directly assigned to me
    from app.models.requirement import Requirement
    assigned_reqs = Requirement.query.filter(
        Requirement.assignee_id == current_user.id,
        Requirement.status.notin_(['done', 'closed']),
    ).all()
    for r in assigned_reqs:
        my_reqs.add(r)
    my_reqs = _urgency_sort(list(my_reqs), limit=30)

    req_days = {}
    for t in my_done:
        for r in t.requirements:
            req_days[r.number] = req_days.get(r.number, 0) + 1

    # Categorize active todos (needed for both GET and POST)
    overdue_todos = [t for t in my_active if t.created_date and t.created_date < monday]
    blocked_todos = [t for t in my_active if t.need_help]

    # Load saved report (if exists)
    saved = PersonalWeekly.query.filter_by(
        user_id=current_user.id, week_start=monday).first()

    report = None
    ai_report = None

    if request.method == 'POST':
        import markdown as md_lib

        # If no data at all, skip AI and show hint
        if not my_done and not my_active and not my_reqs:
            ai_report = '<p>本周暂无工作数据（无已完成或进行中的任务），无法生成周报。</p>'
            ai_report = md_lib.markdown(ai_report)
            report = True
            # Still save
            if saved:
                saved.ai_html = ai_report
            else:
                saved = PersonalWeekly(user_id=current_user.id, week_start=monday,
                                       week_end=sunday, ai_html=ai_report)
                db.session.add(saved)
            db.session.commit()
        else:

            lines = [f'本周（{monday} ~ {sunday}）{current_user.name} 的工作数据：\n']
            normal_active = [t for t in my_active if t not in overdue_todos and t not in blocked_todos]

            if my_done:
                lines.append('本周已完成：')
                for t in my_done:
                    reqs = ', '.join(r.number for r in t.requirements)
                    time_str = f'，用时{t.actual_minutes}分钟' if t.actual_minutes else ''
                    lines.append(f'- ✓ {t.title}（{reqs}{time_str}）')

            if normal_active:
                lines.append('\n进行中：')
                for t in normal_active:
                    reqs = ', '.join(r.number for r in t.requirements)
                    lines.append(f'- ○ {t.title}（{reqs}）')

            if overdue_todos:
                lines.append('\n延期未完成（上周遗留）：')
                for t in overdue_todos:
                    reqs = ', '.join(r.number for r in t.requirements)
                    days = (date.today() - t.created_date).days
                    lines.append(f'- ⚠️ {t.title}（{reqs}，已延期{days}天）')

            if blocked_todos:
                lines.append('\n阻塞中：')
                for t in blocked_todos:
                    reqs = ', '.join(r.number for r in t.requirements)
                    reason = f'，原因：{t.blocked_reason}' if t.blocked_reason else ''
                    lines.append(f'- 🔴 {t.title}（{reqs}{reason}）')

            if my_reqs:
                lines.append('\n参与的需求及状态：')
                for r in my_reqs:
                    due_info = ''
                    if r.due_date:
                        days_left = (r.due_date - date.today()).days
                        due_info = f'，已延期{-days_left}天' if days_left < 0 else f'，剩{days_left}天'
                    child_tag = '，↳子需求' if r.parent_id else ''
                    start_info = f'，开始 {r.start_date.strftime("%m-%d")}' if r.start_date else ''
                    lines.append(f'- {r.title}（{r.status_label}{start_info}{due_info}{child_tag}）')

            # Recurring todo stats this week
            from app.models.recurring_completion import RecurringCompletion
            from app.models.recurring_todo import RecurringTodo
            all_recurring = RecurringTodo.query.filter_by(user_id=current_user.id, is_active=True).all()
            if all_recurring:
                # Count how many times each recurring was due this week and how many were completed
                recurring_due_count = 0
                recurring_done_count = 0
                # Get all completions this week in one query
                week_completions = set()
                for c in RecurringCompletion.query.filter(
                    RecurringCompletion.user_id == current_user.id,
                    RecurringCompletion.recurring_id.in_([r.id for r in all_recurring]),
                    RecurringCompletion.completed_date >= monday,
                    RecurringCompletion.completed_date <= sunday,
                ).all():
                    week_completions.add((c.recurring_id, c.completed_date))

                for d in range((sunday - monday).days + 1):
                    check_date = monday + timedelta(days=d)
                    if check_date > date.today():
                        break
                    for r in all_recurring:
                        is_due = False
                        if r.cycle == 'weekly' and check_date.weekday() == 0:
                            is_due = True
                        elif r.cycle == 'monthly':
                            for p in r.monthly_periods:
                                if check_date.day == r._period_day(p, check_date.year, check_date.month):
                                    is_due = True
                                    break
                        elif r.cycle == 'weekdays' and r.weekdays and str(check_date.weekday()) in r.weekdays.split(','):
                            is_due = True
                        if is_due:
                            recurring_due_count += 1
                            if (r.id, check_date) in week_completions:
                                recurring_done_count += 1

                if recurring_due_count > 0:
                    rate = round(recurring_done_count / recurring_due_count * 100)
                    lines.append('\n周期任务执行情况：')
                    lines.append(f'- 本周到期 {recurring_due_count} 次，完成 {recurring_done_count} 次（完成率 {rate}%）')
                    lines.append(f'- 周期任务共 {len(all_recurring)} 个：' + '、'.join(r.title for r in all_recurring))

            prompt = get_prompt('personal_weekly') + '\n\n' + '\n'.join(lines)
            _, raw = call_ollama(prompt, response_format='text')
            ai_report = raw or 'AI服务暂不可用，正在紧急修复'
            ai_report = md_lib.markdown(ai_report, extensions=['tables'])
            report = True

            # Auto-save
            if saved:
                saved.ai_html = ai_report
            else:
                saved = PersonalWeekly(
                    user_id=current_user.id, week_start=monday,
                    week_end=sunday, ai_html=ai_report)
                db.session.add(saved)
            db.session.commit()

    elif saved and saved.ai_html:
        # Load previously saved report
        ai_report = saved.ai_html
        report = True

    # Calculate totals
    total_focus = sum(t.actual_minutes or 0 for t in my_done)
    reviewer_name = get_reviewer(current_user)

    # User's open risks (assigned as owner or tracker)
    from app.models.risk import Risk
    my_open_risks = Risk.query.filter(
        Risk.deleted_at.is_(None), Risk.status == 'open',
        db.or_(Risk.owner_id == current_user.id, Risk.tracker_id == current_user.id),
    ).order_by(Risk.due_date).all()

    # User's open requirements (assigned to me, not done/closed)
    from app.models.requirement import Requirement
    my_open_reqs = Requirement.query.filter(
        Requirement.assignee_id == current_user.id,
        Requirement.status.notin_(['done', 'closed']),
    ).order_by(Requirement.due_date).all()

    from app.utils.recipients import compute_personal_recipients
    _def_to, _def_cc = compute_personal_recipients(current_user)

    # ---- Week-over-week delta (last week vs this week) ----
    prev_mon, prev_sun = week_range(offset - 1)
    _prev_done_filter = db.and_(
        Todo.status == 'done',
        db.or_(
            db.and_(Todo.done_date >= prev_mon, Todo.done_date <= prev_sun),
            db.and_(Todo.done_date.is_(None), Todo.created_date >= prev_mon, Todo.created_date <= prev_sun),
        ),
    )
    from sqlalchemy import func as _fn2, case as _case2
    _prev = db.session.query(
        # prev_done_cnt
        _fn2.sum(_case2((_prev_done_filter, 1), else_=0)),
        # prev_active_cnt
        _fn2.sum(_case2((db.and_(Todo.status == 'todo', Todo.created_date <= prev_sun), 1), else_=0)),
        # prev_overdue_cnt
        _fn2.sum(_case2((db.and_(Todo.status == 'todo', Todo.created_date < prev_mon), 1), else_=0)),
        # prev_focus (sum of actual_minutes for done todos in prev week)
        _fn2.coalesce(_fn2.sum(_case2(
            (db.and_(_prev_done_filter, Todo.actual_minutes > 0), Todo.actual_minutes),
            else_=0,
        )), 0),
    ).filter(Todo.user_id == current_user.id).first()
    prev_done_cnt = int(_prev[0] or 0)
    prev_active_cnt = int(_prev[1] or 0)
    prev_overdue_cnt = int(_prev[2] or 0)
    prev_focus = int(_prev[3] or 0)
    deltas = {
        'done': len(my_done) - prev_done_cnt,
        'active': len(my_active) - prev_active_cnt,
        'overdue': len(overdue_todos) - prev_overdue_cnt,
        'focus': total_focus - prev_focus,
    }

    # ---- Group by project, sort by max overdue days (most overdue first) ----
    from collections import defaultdict as _ddict
    _today = date.today()

    def _group_by_project(items, get_project, get_overdue_days):
        groups = _ddict(list)
        for item in items:
            pname = get_project(item)
            groups[pname].append(item)
        # Sort projects: max overdue days desc, then project name
        def _proj_sort_key(pname):
            max_od = max((get_overdue_days(i) for i in groups[pname]), default=0)
            return (-max_od, pname)
        sorted_groups = []
        for pname in sorted(groups.keys(), key=_proj_sort_key):
            sorted_groups.append((pname, groups[pname]))
        return sorted_groups

    reqs_by_project = _group_by_project(
        my_reqs,
        lambda r: (r.project.name[:6] if r.project else '-'),
        lambda r: ((_today - r.due_date).days if r.due_date and r.due_date < _today and r.status not in ('done', 'closed') else 0),
    )
    risks_by_project = _group_by_project(
        my_open_risks,
        lambda r: (r.project.name[:6] if r.project else '-'),
        lambda r: ((_today - r.due_date).days if r.due_date and r.due_date < _today else 0),
    )

    return render_template('dashboard/my_weekly.html',
        my_done=my_done, my_active=my_active, my_reqs=my_reqs,
        reqs_by_project=reqs_by_project, risks_by_project=risks_by_project,
        req_days=req_days, report=report, ai_report=ai_report,
        overdue_todos=overdue_todos, blocked_todos=blocked_todos,
        total_focus_min=total_focus, reviewer=reviewer_name,
        my_open_risks=my_open_risks,
        deltas=deltas,
        today=date.today(),
        monday=monday, sunday=sunday, offset=offset,
        default_to=_def_to, default_cc=_def_cc,
    )


# ---- Resource allocation map ----

@dashboard_bp.route('/resource-map')
@login_required
def resource_map():
    from collections import defaultdict

    period = request.args.get('period', 'week')
    mode = request.args.get('mode', 'by_person')  # by_person or by_project
    week_offset = request.args.get('week', 0, type=int)

    today = date.today()
    if period == '3month':
        start = today - timedelta(days=90)
        end = today
    elif period == 'last_month':
        first_this_month = today.replace(day=1)
        end = first_this_month - timedelta(days=1)
        start = end.replace(day=1)
    elif period == 'month':
        start = today.replace(day=1)
        end = today
    elif period == 'last_week':
        start, end = week_range(-1)
    else:
        start, end = week_range(week_offset)
    label = f'{start.strftime("%Y-%m-%d")} ~ {end.strftime("%Y-%m-%d")}'

    cur_group = request.args.get('group', '')
    cur_project_filter = request.args.get('project', 0, type=int)
    hidden_groups = {g.name for g in Group.query.filter_by(is_hidden=True).all()}
    visible_groups = [g.name for g in Group.query.filter_by(is_hidden=False).order_by(Group.name).all()]
    users = [u for u in User.query.filter_by(is_active=True).order_by(User.group, User.name).all()
             if u.group not in hidden_groups]
    if cur_group:
        users = [u for u in users if u.group == cur_group]
    user_ids = [u.id for u in users]

    # All todos in the period (by created_date, which always exists)
    todos = Todo.query.filter(
        Todo.user_id.in_(user_ids),
        Todo.created_date >= start, Todo.created_date <= end,
    ).options(joinedload(Todo.requirements)).all()

    # Per day: count todos per project, then split the day proportionally
    # e.g. day has 5 todos for projA + 1 for projB → projA=5/6天, projB=1/6天
    user_date_proj_count = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for t in todos:
        work_date = t.done_date or t.created_date
        if not work_date:
            continue
        for r in t.requirements:
            if r.project_id:
                user_date_proj_count[t.user_id][work_date][r.project_id] += 1

    user_project_days = defaultdict(float)
    for uid, dates in user_date_proj_count.items():
        for _dt, proj_counts in dates.items():
            day_total = sum(proj_counts.values())
            if day_total <= 0:
                continue
            for pid, cnt in proj_counts.items():
                user_project_days[(uid, pid)] += cnt / day_total

    project_ids = sorted(set(pid for (_, pid) in user_project_days))
    projects = {p.id: p for p in Project.query.filter(Project.id.in_(project_ids)).all()} if project_ids else {}

    # Filter out hidden projects for non-managers
    _hset = set(g.hidden_pids)
    if _hset:
        project_ids = [pid for pid in project_ids if pid not in _hset]
        projects = {pid: p for pid, p in projects.items() if pid not in _hset}
        user_project_days = {k: v for k, v in user_project_days.items() if k[1] not in _hset}

    # Per-user total days
    user_total = defaultdict(float)
    for (uid, _pid), d in user_project_days.items():
        user_total[uid] += d

    # Load expected_ratio from ProjectMember
    member_map = {}
    if project_ids:
        members = ProjectMember.query.filter(
            ProjectMember.project_id.in_(project_ids),
            ProjectMember.user_id.isnot(None)).all()
        for m in members:
            member_map[(m.user_id, m.project_id)] = m

    # Mode 1 (by_person): flat rows — one row per (person, project)
    flat_rows = []
    for u in users:
        t = user_total.get(u.id, 0)
        if t <= 0:
            continue
        for pid in project_ids:
            d = user_project_days.get((u.id, pid), 0)
            if d <= 0:
                continue
            ratio = round(d / t * 100) if t else 0
            member = member_map.get((u.id, pid))
            flat_rows.append({
                'user': u,
                'project': projects.get(pid),
                'project_id': pid,
                'days': round(d, 1),
                'ratio': ratio,
                'expected_ratio': member.expected_ratio if member else None,
            })

    # Mode 2 (by_project): flat rows — one row per (project, person)
    proj_flat_rows = []
    for pid in project_ids:
        p = projects.get(pid)
        if not p:
            continue
        proj_total = sum(user_project_days.get((u.id, pid), 0) for u in users)
        if proj_total <= 0:
            continue
        for u in users:
            d = user_project_days.get((u.id, pid), 0)
            if d <= 0:
                continue
            ratio = round(d / proj_total * 100) if proj_total else 0
            proj_flat_rows.append({
                'project': p,
                'project_id': pid,
                'user': u,
                'days': round(d, 1),
                'proj_total': round(proj_total, 1),
                'ratio': ratio,
            })

    # Apply project filter
    if cur_project_filter:
        flat_rows = [r for r in flat_rows if r['project_id'] == cur_project_filter]
        proj_flat_rows = [r for r in proj_flat_rows if r['project_id'] == cur_project_filter]

    is_pm = current_user.is_admin or current_user.has_role('PM', 'PL', 'FO', 'LM', 'XM', 'HR')
    return render_template('dashboard/resource_map.html',
        flat_rows=flat_rows, proj_flat_rows=proj_flat_rows,
        projects=projects, project_ids=project_ids, users=users,
        period=period, mode=mode, label=label, offset=week_offset,
        is_pm=is_pm, groups=visible_groups, cur_group=cur_group,
        cur_project_filter=cur_project_filter,
    )


@dashboard_bp.route('/resource-map/export')
@login_required
def resource_map_export():
    """Export resource map as CSV."""
    import csv
    import io
    period = request.args.get('period', 'week')
    mode = request.args.get('mode', 'by_person')
    week_offset = request.args.get('week', 0, type=int)

    # Reuse the same logic — call resource_map internally
    with current_app.test_request_context(f'/dashboard/resource-map?period={period}&week={week_offset}&mode={mode}'):
        from flask_login import login_user
        login_user(current_user)

    # Simpler: just query directly
    from collections import defaultdict
    today = date.today()
    if period == '3month':
        start = today - timedelta(days=90)
        end = today
    elif period == 'last_month':
        first_this_month = today.replace(day=1)
        end = first_this_month - timedelta(days=1)
        start = end.replace(day=1)
    elif period == 'month':
        start = today.replace(day=1)
        end = today
    elif period == 'last_week':
        start, end = week_range(-1)
    else:
        start, end = week_range(week_offset)

    _hg = {g.name for g in Group.query.filter_by(is_hidden=True).all()}
    users_q = [u for u in User.query.filter_by(is_active=True).order_by(User.group, User.name).all()
               if u.group not in _hg]
    user_ids = [u.id for u in users_q]
    todos = Todo.query.filter(
        Todo.user_id.in_(user_ids),
        Todo.created_date >= start, Todo.created_date <= end,
    ).options(joinedload(Todo.requirements)).all()

    user_date_proj_count = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    for t in todos:
        work_date = t.done_date or t.created_date
        if not work_date:
            continue
        for r in t.requirements:
            if r.project_id:
                user_date_proj_count[t.user_id][work_date][r.project_id] += 1

    user_project_days = defaultdict(float)
    for uid, dates_data in user_date_proj_count.items():
        for _dt, proj_counts in dates_data.items():
            day_total = sum(proj_counts.values())
            if day_total <= 0:
                continue
            for pid, cnt in proj_counts.items():
                user_project_days[(uid, pid)] += cnt / day_total

    project_ids = sorted(set(pid for (_, pid) in user_project_days))
    projects = {p.id: p for p in Project.query.filter(Project.id.in_(project_ids)).all()} if project_ids else {}

    # Filter out hidden projects for non-managers
    _hset = set(g.hidden_pids)
    if _hset:
        project_ids = [pid for pid in project_ids if pid not in _hset]
        projects = {pid: p for pid, p in projects.items() if pid not in _hset}
        user_project_days = {k: v for k, v in user_project_days.items() if k[1] not in _hset}

    user_total = defaultdict(float)
    for (uid, _pid), d in user_project_days.items():
        user_total[uid] += d

    buf = io.StringIO()
    writer = csv.writer(buf)
    if mode == 'by_person':
        writer.writerow(['姓名', '工号', '项目', '投入天数', '实际比例%', '预期比例%'])
        for u in users_q:
            t = user_total.get(u.id, 0)
            if t <= 0:
                continue
            for pid in project_ids:
                d = user_project_days.get((u.id, pid), 0)
                if d <= 0:
                    continue
                ratio = round(d / t * 100) if t else 0
                member = ProjectMember.query.filter_by(user_id=u.id, project_id=pid).first()
                exp = member.expected_ratio if member and member.expected_ratio else ''
                writer.writerow([u.name, u.employee_id, projects.get(pid, type('', (), {'name': '-'})).name,
                                 round(d, 1), ratio, exp])
    else:
        writer.writerow(['项目', '项目合计', '姓名', '工号', '投入天数', '占比%'])
        for pid in project_ids:
            p = projects.get(pid)
            if not p:
                continue
            proj_total = sum(user_project_days.get((u.id, pid), 0) for u in users_q)
            if proj_total <= 0:
                continue
            for u in users_q:
                d = user_project_days.get((u.id, pid), 0)
                if d <= 0:
                    continue
                ratio = round(d / proj_total * 100)
                writer.writerow([p.name, round(proj_total, 1), u.name, u.employee_id, round(d, 1), ratio])

    output = buf.getvalue().encode('utf-8-sig')
    return send_file(io.BytesIO(output), download_name=f'人力投入_{start}_{end}.csv',
                     as_attachment=True, mimetype='text/csv')


@dashboard_bp.route('/resource-map/expected-ratio', methods=['POST'])
@login_required
def save_expected_ratio():

    if not (current_user.is_admin or current_user.has_role('PM', 'PL', 'FO', 'LM', 'XM', 'HR')):
        return jsonify(ok=False, msg='无权限'), 403
    uid = request.form.get('user_id', type=int)
    pid = request.form.get('project_id', type=int)
    ratio = request.form.get('ratio', type=int)
    if not uid or not pid:
        return jsonify(ok=False), 400
    member = ProjectMember.query.filter_by(user_id=uid, project_id=pid).first()
    if not member:
        member = ProjectMember(user_id=uid, project_id=pid, project_role='DEV')
        db.session.add(member)
    member.expected_ratio = ratio if ratio and ratio > 0 else None
    db.session.commit()
    return jsonify(ok=True)


# ---- Compliance ----

@dashboard_bp.route('/compliance')
@login_required
def compliance():
    """合规专区：汇总合规相关功能入口和状态。"""
    from app.models.audit import AuditLog
    from app.models.risk import Risk

    # 审计日志近30天统计
    _30d = date.today() - timedelta(days=30)
    audit_count = AuditLog.query.filter(AuditLog.created_at >= _30d).count()

    # 未闭环风险
    open_risks = Risk.query.filter_by(status='open').filter(Risk.deleted_at.is_(None)).count()
    overdue_risks = Risk.query.filter(
        Risk.status == 'open', Risk.deleted_at.is_(None),
        Risk.due_date < date.today(), Risk.due_date.isnot(None)
    ).count()

    # 权限申请待审批
    from app.models.knowledge import PermissionApplication
    pending_perms = PermissionApplication.query.filter_by(status='pending').count()

    from app.models.site_setting import SiteSetting

    # 入职引导：签署后不再显示
    _onboarded = set(SiteSetting.get('onboarded_users', '').split(','))
    show_onboarding = str(current_user.id) not in _onboarded

    # 离职入口：仅管理层在后台对具体人员开启后才显示
    _offboard_uids = SiteSetting.get('offboarding_users', '')
    show_offboarding = str(current_user.id) in _offboard_uids.split(',') if _offboard_uids else False

    # 入职必读必学必考配置
    import json as _json
    _onboard_cfg_raw = SiteSetting.get('onboard_config', '')
    try:
        onboard_config = _json.loads(_onboard_cfg_raw) if _onboard_cfg_raw else {}
    except Exception:
        onboard_config = {}

    # 合规考试统计
    _q_raw = SiteSetting.get('compliance_questions', '')
    exam_question_count = len(_json.loads(_q_raw)) if _q_raw else 0
    _passed_raw = SiteSetting.get('compliance_exam_passed', '')
    exam_passed_uids = set(_passed_raw.split(',')) if _passed_raw else set()
    exam_passed_uids.discard('')
    exam_passed = str(current_user.id) in exam_passed_uids

    # 考勤诚信承诺
    _att_signed = set(SiteSetting.get('attendance_signed_users', '').split(','))
    _att_signed.discard('')
    attendance_signed = str(current_user.id) in _att_signed

    # 考试诚信承诺
    _exam_integrity = set(SiteSetting.get('exam_integrity_signed', '').split(','))
    _exam_integrity.discard('')
    exam_integrity_signed = str(current_user.id) in _exam_integrity

    return render_template('dashboard/compliance.html',
        audit_count=audit_count, open_risks=open_risks,
        overdue_risks=overdue_risks, pending_perms=pending_perms,
        show_onboarding=show_onboarding, show_offboarding=show_offboarding,
        onboard_config=onboard_config,
        exam_question_count=exam_question_count, exam_passed=exam_passed,
        attendance_signed=attendance_signed,
        exam_integrity_signed=exam_integrity_signed)


@dashboard_bp.route('/compliance/onboard-sign', methods=['POST'])
@login_required
def compliance_onboard_sign():
    """签署入职引导，标记为已完成。"""
    from app.models.site_setting import SiteSetting
    raw = SiteSetting.get('onboarded_users', '')
    uids = set(raw.split(',')) if raw else set()
    uids.discard('')
    uids.add(str(current_user.id))
    SiteSetting.set('onboarded_users', ','.join(sorted(uids)))
    return jsonify(ok=True)


@dashboard_bp.route('/compliance/attendance-sign', methods=['POST'])
@login_required
def compliance_attendance_sign():
    """签署考勤诚信承诺。"""
    from app.models.site_setting import SiteSetting
    raw = SiteSetting.get('attendance_signed_users', '')
    uids = set(raw.split(',')) if raw else set()
    uids.discard('')
    uids.add(str(current_user.id))
    SiteSetting.set('attendance_signed_users', ','.join(sorted(uids)))
    return jsonify(ok=True)


@dashboard_bp.route('/compliance/exam-integrity-sign', methods=['POST'])
@login_required
def compliance_exam_integrity_sign():
    """签署考试诚信承诺。"""
    from app.models.site_setting import SiteSetting
    raw = SiteSetting.get('exam_integrity_signed', '')
    uids = set(raw.split(',')) if raw else set()
    uids.discard('')
    uids.add(str(current_user.id))
    SiteSetting.set('exam_integrity_signed', ','.join(sorted(uids)))
    return jsonify(ok=True)


@dashboard_bp.route('/compliance/onboard-config', methods=['POST'])
@login_required
def compliance_onboard_config():
    """保存入职必读必学必考配置（所有人可添加，删除由前端控制仅管理层）。"""
    import json as _json
    from app.models.site_setting import SiteSetting
    data = request.get_json() or {}
    SiteSetting.set('onboard_config', _json.dumps(data, ensure_ascii=False))
    return jsonify(ok=True)


@dashboard_bp.route('/compliance/exam', methods=['GET'])
@login_required
def compliance_exam_page():
    """合规考试页面：随机抽题。"""
    import json as _json, random
    from app.models.site_setting import SiteSetting
    raw = SiteSetting.get('compliance_questions', '')
    questions = _json.loads(raw) if raw else []
    if not questions:
        flash('题库为空，请联系管理员', 'warning')
        return redirect(url_for('dashboard.compliance'))
    # 随机抽取（最多20题）
    pool = list(questions)
    random.shuffle(pool)
    exam = pool[:min(20, len(pool))]
    return render_template('dashboard/compliance_exam.html', questions=exam)


@dashboard_bp.route('/compliance/exam-submit', methods=['POST'])
@login_required
def compliance_exam_submit():
    """提交合规考试答案。"""
    import json as _json
    from app.models.site_setting import SiteSetting
    raw = SiteSetting.get('compliance_questions', '')
    all_q = _json.loads(raw) if raw else []
    data = request.get_json() or {}
    answers = data.get('answers', {})  # { stem: chosen_key }
    # 对照批改
    q_map = {q['stem']: q['answer'] for q in all_q}
    total = len(answers)
    correct = sum(1 for stem, ans in answers.items() if q_map.get(stem) == ans)
    passed = total > 0 and (correct / total) >= 0.8
    if passed:
        prev = SiteSetting.get('compliance_exam_passed', '')
        uids = set(prev.split(',')) if prev else set()
        uids.discard('')
        uids.add(str(current_user.id))
        SiteSetting.set('compliance_exam_passed', ','.join(sorted(uids)))
    return jsonify(ok=True, total=total, correct=correct, passed=passed)


# ---- Emotion prediction ----

def _emotion_guard():
    """情绪预测：仅管理层+私密模式(eye开)可访问。"""
    if not current_user.is_team_manager:
        abort(403)
    if request.cookies.get('mgr_view') != '1':
        abort(403)
    if not (current_user.is_admin or current_user.has_role('PL', 'LM', 'XM', 'HR')):
        abort(403)

@dashboard_bp.route('/emotion')
@login_required
def emotion_predict():
    _emotion_guard()
    from app.models.emotion import EmotionRecord
    visible_groups = [g.name for g in Group.query.filter_by(is_hidden=False).order_by(Group.name).all()]
    is_senior = current_user.is_admin or current_user.has_role('LM', 'XM', 'HR')
    # PL 默认看自己组，XM/HR/LM 默认看全部
    if 'group' in request.args:
        cur_group = request.args.get('group', '')
    else:
        cur_group = '' if is_senior else (current_user.group or '')

    # Load saved records grouped by date
    risk_order = db.case(
        (EmotionRecord.risk_level == 'high', 0),
        (EmotionRecord.risk_level == 'medium', 1),
        else_=2
    )
    q = EmotionRecord.query
    if cur_group:
        q = q.filter_by(group=cur_group)
    records = q.order_by(EmotionRecord.scan_date.desc(), risk_order).all()
    dates = sorted(set(r.scan_date for r in records), reverse=True)
    grouped = {}
    for d in dates:
        grouped[d] = [r for r in records if r.scan_date == d]

    # 1v1 覆盖率：每人最近一次记录 + 距今天数
    hidden_groups = {g.name for g in Group.query.filter_by(is_hidden=True).all()}
    all_users = User.query.filter_by(is_active=True).order_by(User.group, User.name).all()
    all_users = [u for u in all_users if u.group not in hidden_groups]
    if cur_group:
        all_users = [u for u in all_users if u.group == cur_group]

    # Build per-member latest record (across all records, not just filtered)
    all_records = EmotionRecord.query.order_by(EmotionRecord.scan_date.desc()).all()
    member_latest = {}  # member_name → latest EmotionRecord
    member_history = {}  # member_name → [records oldest→newest]
    for r in all_records:
        if r.member_name not in member_latest:
            member_latest[r.member_name] = r
        member_history.setdefault(r.member_name, []).append(r)
    # Reverse history to oldest-first
    for k in member_history:
        member_history[k] = list(reversed(member_history[k]))

    today_ = date.today()
    coverage = []
    for u in all_users:
        latest = member_latest.get(u.name)
        days_since = (today_ - latest.scan_date).days if latest else None
        history = member_history.get(u.name, [])
        coverage.append({
            'user': u,
            'latest': latest,
            'days_since': days_since,
            'history': history,
            'overdue': days_since is None or days_since >= 60,
            'warning': days_since is not None and 30 <= days_since < 60,
        })
    # Sort: overdue first (no record → top, then by days_since desc)
    coverage.sort(key=lambda c: (0 if c['days_since'] is None else 1, -(c['days_since'] or 999)))

    # 各组汇总统计（XM/HR/LM 视角）
    group_stats = {}
    if is_senior:
        from collections import defaultdict
        _gs = defaultdict(lambda: {'total': 0, 'overdue': 0, 'warning': 0, 'ok': 0})
        # Use all users (not filtered by cur_group) to compute cross-group stats
        _all_users_for_stats = [u for u in User.query.filter_by(is_active=True).all()
                                if u.group not in hidden_groups and u.group]
        for u in _all_users_for_stats:
            g_name = u.group
            _gs[g_name]['total'] += 1
            latest = member_latest.get(u.name)
            ds = (today_ - latest.scan_date).days if latest else None
            if ds is None or ds >= 60:
                _gs[g_name]['overdue'] += 1
            elif ds >= 30:
                _gs[g_name]['warning'] += 1
            else:
                _gs[g_name]['ok'] += 1
        group_stats = {k: dict(v) for k, v in sorted(_gs.items())}

    # 1v1 谈话模版（后台可配置）
    from app.models.site_setting import SiteSetting
    talk_tpl_raw = SiteSetting.get('emotion_talk_template', '')
    talk_items = [line.strip() for line in talk_tpl_raw.strip().splitlines() if line.strip()] if talk_tpl_raw else []

    return render_template('dashboard/emotion.html', grouped=grouped, dates=dates,
                           today=today_, groups=visible_groups, cur_group=cur_group,
                           coverage=coverage, talk_items=talk_items,
                           is_senior=is_senior, group_stats=group_stats)


@dashboard_bp.route('/emotion/analyze', methods=['POST'])
@login_required
def emotion_analyze():
    """AI analyzes team emotion and attrition risk."""
    _emotion_guard()

    cur_group = request.args.get('group', '') or request.form.get('group', '')
    hidden_groups = {g.name for g in Group.query.filter_by(is_hidden=True).all()}
    users = [u for u in User.query.filter_by(is_active=True).order_by(User.name).all()
             if u.group not in hidden_groups]
    if cur_group:
        users = [u for u in users if u.group == cur_group]
    today = date.today()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    lines = [f'分析日期：{today}\n']
    lines.append('团队成员近期工作数据：\n')

    for u in users:
        # Last 7 days
        done_7d = Todo.query.filter(
            Todo.user_id == u.id, Todo.done_date >= week_ago, Todo.done_date <= today).count()
        # Last 30 days
        done_30d = Todo.query.filter(
            Todo.user_id == u.id, Todo.done_date >= month_ago, Todo.done_date <= today).count()
        # Active (unfinished)
        active = Todo.query.filter_by(user_id=u.id, status='todo').count()
        # Blocked
        blocked = Todo.query.filter(
            Todo.user_id == u.id, Todo.status == 'todo', Todo.need_help == True).count()
        # Help given (source='help', this user as helper)
        help_given = Todo.query.filter(
            Todo.user_id == u.id, Todo.source == 'help', Todo.created_date >= month_ago).count()
        # Focus time
        focus = db.session.query(db.func.sum(Todo.actual_minutes)).filter(
            Todo.user_id == u.id, Todo.created_date >= month_ago).scalar() or 0
        # Last active date
        last_done = db.session.query(db.func.max(Todo.done_date)).filter(
            Todo.user_id == u.id).scalar()
        last_str = str(last_done) if last_done else '无记录'
        # Daily avg (30d)
        daily_avg = round(done_30d / 30, 1) if done_30d else 0

        lines.append(
            f'- {u.name}（{u.group or ""}）：\n'
            f'  近7天完成 {done_7d} 个 | 近30天完成 {done_30d} 个（日均 {daily_avg}）\n'
            f'  进行中 {active} 个 | 阻塞 {blocked} 个 | 协助他人 {help_given} 次\n'
            f'  番茄钟 {focus} 分钟 | 最后完成日期 {last_str}')

    # 补充近2个月的1v1聊天记录
    from app.models.emotion import EmotionRecord
    two_months_ago = today - timedelta(days=60)
    _rq = EmotionRecord.query.filter(EmotionRecord.scan_date >= two_months_ago)
    if cur_group:
        _rq = _rq.filter_by(group=cur_group)
    recent_records = _rq.order_by(EmotionRecord.scan_date.desc()).all()
    if recent_records:
        lines.append('\n\n近2个月1v1谈话记录：\n')
        for r in recent_records:
            rec_line = f'- {r.member_name}（{r.group or ""}）{r.scan_date}：状态={r.status}，风险={r.risk_level}'
            if r.suggestion:
                # 过滤掉"未提及"的观察项，减少 token
                useful = [l for l in r.suggestion.split('\n') if l.strip() and '未提及' not in l]
                if useful:
                    rec_line += f'，记录={"; ".join(useful)}'
            if r.signals_list:
                rec_line += f'，信号={"; ".join(r.signals_list)}'
            if r.comments:
                for c in r.comments[:3]:  # 最多3条跟进
                    rec_line += f'，跟进({c.user.name})：{c.content}'
            lines.append(rec_line)

    prompt = get_prompt('emotion_predict') + '\n\n' + '\n'.join(lines)
    result, raw = call_ollama(prompt)

    if isinstance(result, list):
        return jsonify(ok=True, members=result)
    return jsonify(ok=False, raw=raw or 'AI服务暂不可用，正在紧急修复')


@dashboard_bp.route('/emotion/save', methods=['POST'])
@login_required
def emotion_save():
    """Save AI emotion analysis results."""
    _emotion_guard()
    import json as json_lib

    from app.models.emotion import EmotionRecord
    data = request.get_json() or {}
    members = data.get('members', [])
    today = date.today()
    # 只删今天的 AI 生成记录（signals 以 "__ai__" 开头），保留手动 1v1 记录
    for old in EmotionRecord.query.filter_by(scan_date=today).all():
        if old.signals and old.signals.startswith('["__ai__"'):
            db.session.delete(old)
    for m in members:
        # AI 记录在 signals 首位加 "__ai__" 标记
        ai_signals = ['__ai__'] + (m.get('signals') or [])
        db.session.add(EmotionRecord(
            scan_date=today,
            member_name=m.get('name', ''),
            group=m.get('group', ''),
            status=m.get('status', '正常'),
            risk_level=m.get('risk_level', 'low'),
            signals=json_lib.dumps(ai_signals, ensure_ascii=False),
            suggestion=m.get('suggestion', ''),
            created_by=current_user.id,
        ))
    db.session.commit()
    return jsonify(ok=True, count=len(members))


@dashboard_bp.route('/emotion/delete-record/<int:record_id>', methods=['POST'])
@login_required
def emotion_delete_record(record_id):
    """Delete a single emotion record."""
    _emotion_guard()
    from app.models.emotion import EmotionRecord
    r = db.session.get(EmotionRecord, record_id)
    if r:
        db.session.delete(r)
        db.session.commit()
    return jsonify(ok=True)


@dashboard_bp.route('/emotion/delete/<scan_date>', methods=['POST'])
@login_required
def emotion_delete(scan_date):
    """Delete all emotion records for a specific date."""
    _emotion_guard()
    from app.models.emotion import EmotionRecord
    _ = EmotionRecord.query.filter_by(scan_date=scan_date).delete()
    db.session.commit()
    flash(f'已删除 {scan_date} 的记录', 'success')
    return redirect(url_for('dashboard.emotion_predict'))


@dashboard_bp.route('/emotion/comment/<int:record_id>', methods=['POST'])
@login_required
def emotion_comment(record_id):
    """Add comment to an emotion record. Supports #comment and @person for todo."""
    _emotion_guard()
    import re

    from app.models.emotion import EmotionComment, EmotionRecord
    from app.models.todo import Todo, TodoItem

    record = db.get_or_404(EmotionRecord, record_id)
    content = (request.form.get('content') or '').strip()[:500]
    if not content:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify(ok=False, msg='内容不能为空')
        return redirect(url_for('dashboard.emotion_predict'))

    # Save comment
    comment = EmotionComment(record_id=record.id, user_id=current_user.id, content=content)
    db.session.add(comment)

    # Check for @mention — create a follow-up todo
    todo_created = None
    at_match = re.search(r'@(\S+)', content)
    if at_match:
        target_name = at_match.group(1)
        target_user = User.query.filter(
            db.or_(User.name == target_name, User.pinyin.ilike(f'{target_name}%'))
        ).filter_by(is_active=True).first()
        if target_user:
            clean_content = re.sub(r'@\S+', '', content).strip()
            todo_title = f'[情绪跟进] {record.member_name}：{clean_content[:50]}'
            todo = Todo(user_id=target_user.id, title=todo_title,
                        due_date=date.today() + timedelta(days=7), category='team', source='help')
            todo.items.append(TodoItem(title=todo_title, sort_order=0))
            db.session.add(todo)
            todo_created = target_user.name

    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        from datetime import datetime as _dt
        return jsonify(ok=True, content=content,
                       user=current_user.name,
                       time=_dt.now().strftime('%m-%d %H:%M'),
                       todo_created=todo_created)
    return redirect(url_for('dashboard.emotion_predict'))


@dashboard_bp.route('/emotion/add-record', methods=['POST'])
@login_required
def emotion_add_record():
    """手动添加1v1聊天记录，复用 EmotionRecord 模型。"""
    _emotion_guard()
    import json as json_lib

    from app.models.emotion import EmotionRecord
    member_name = request.form.get('member_name', '').strip()
    if not member_name:
        return jsonify(ok=False, msg='请选择成员')
    user_obj = User.query.filter_by(name=member_name, is_active=True).first()
    member_group = user_obj.group if user_obj else ''
    status = request.form.get('status', '正常')
    risk_level = request.form.get('risk_level', 'low')
    signals_raw = request.form.get('signals', '').strip()
    signals = [s.strip() for s in signals_raw.split('\n') if s.strip()] if signals_raw else []
    suggestion = request.form.get('suggestion', '').strip()
    record_date = request.form.get('record_date', '')
    try:
        from datetime import datetime as _dt
        scan_date = _dt.strptime(record_date, '%Y-%m-%d').date() if record_date else date.today()
    except ValueError:
        scan_date = date.today()

    db.session.add(EmotionRecord(
        scan_date=scan_date,
        member_name=member_name,
        group=member_group,
        status=status,
        risk_level=risk_level,
        signals=json_lib.dumps(signals, ensure_ascii=False) if signals else None,
        suggestion=suggestion or None,
        created_by=current_user.id,
    ))
    db.session.commit()
    flash(f'已记录 {member_name} 的1v1谈话', 'success')
    return redirect(url_for('dashboard.emotion_predict', group=member_group or ''))
